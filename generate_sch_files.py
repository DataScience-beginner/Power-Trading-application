#!/usr/bin/env python3
"""
Generate SCH files to match the DOR files we uploaded
For each DOR file with date X, create an SCH file with date X+1
"""

from pathlib import Path
import shutil
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

def generate_sch_files():
    """Generate SCH files matching DOR files"""
    
    # Source directories
    dor_dir = Path("Data/mock_reports_real_format")
    sch_template = Path("Data/IEX260114SCH_NPT0019_TN0_Grasim_Industries_Limited.xlsx")
    output_dir = Path("Data/mock_reports_sch")
    output_dir.mkdir(exist_ok=True)
    
    if not sch_template.exists():
        print(f"❌ SCH template not found: {sch_template}")
        return
    
    # Get all DOR files
    dor_files = list(dor_dir.glob("*DOR*.xlsx"))
    print(f"Found {len(dor_files)} DOR files\n")
    
    sch_count = 0
    
    for dor_file in dor_files:
        # Parse DOR filename: DAM_IEX110126DOR_NPT0001_MH0_Tata_Power_Limited.xlsx
        # Extract: date (110126), market (DAM), entity (NPT0001), state (MH0), name (Tata Power Limited)
        
        match = re.search(r'(GDAM|DAM|RTM)_IEX(\d{6})DOR_([^_]+)_([^_]+)_(.+)\.xlsx', dor_file.name)
        if not match:
            print(f"⚠️  Skipping {dor_file.name} - can't parse filename")
            continue
        
        market, date_str, entity_id, state_code, company_name = match.groups()
        
        # Parse date and add 1 day for SCH
        try:
            dor_date = datetime.strptime(date_str, "%d%m%y")
            sch_date = dor_date + timedelta(days=1)
            sch_date_str = sch_date.strftime("%d%m%y")
        except Exception as e:
            print(f"⚠️  Skipping {dor_file.name} - can't parse date: {e}")
            continue
        
        # Create SCH filename: IEX020126SCH_NPT0001_MH0_Tata_Power_Limited.xlsx
        sch_filename = f"IEX{sch_date_str}SCH_{entity_id}_{state_code}_{company_name}.xlsx"
        sch_path = output_dir / sch_filename
        
        # Copy template
        shutil.copy(sch_template, sch_path)
        
        # Modify Excel to set correct entity_id and date
        try:
            wb = load_workbook(sch_path)
            ws = wb.active
            
            # Update entity_id (Row 6, Column C typically, but let's check the template structure)
            # For now, just update the delivery date (Row 8, Column I)
            # Note: Row indexing in openpyxl is 1-based
            
            # Set delivery date
            ws[f'I8'] = sch_date  # Delivery Date
            
            # Set entity ID if we can find the cell
            # Typically entity_id is in row 6, column C
            if ws['C6'].value:
                ws['C6'] = entity_id
            
            wb.save(sch_path)
            sch_count += 1
            
            if sch_count % 10 == 0:
                print(f"   Generated {sch_count} SCH files...")
            
        except Exception as e:
            print(f"❌ Error modifying {sch_filename}: {e}")
    
    print(f"\n✅ Generated {sch_count} SCH files in {output_dir}")
    return sch_count

if __name__ == "__main__":
    generate_sch_files()
