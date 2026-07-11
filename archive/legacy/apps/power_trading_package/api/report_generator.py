"""
Report Generation Module
Generates PDF and Excel reports from trading data
"""

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime
from typing import List, Dict, Any

def generate_daily_trading_pdf(transactions: List[Dict], date: str) -> BytesIO:
    """Generate PDF report for daily trading"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"Daily Trading Report - {date}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary statistics
    total_volume = sum(t.get('quantity_mw', 0) for t in transactions)
    total_amount = sum(t.get('amount', 0) for t in transactions)
    buy_count = len([t for t in transactions if t.get('transaction_type') == 'buy'])
    sell_count = len([t for t in transactions if t.get('transaction_type') == 'sell'])
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', str(len(transactions))],
        ['Buy Transactions', str(buy_count)],
        ['Sell Transactions', str(sell_count)],
        ['Total Volume', f"{total_volume:.2f} MW"],
        ['Total Amount', f"₹{total_amount:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Transaction details (first 50 transactions)
    elements.append(Paragraph("Transaction Details (Top 50)", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    trans_data = [['Time', 'Type', 'Quantity (MW)', 'Rate (₹/MWh)', 'Amount (₹)']]
    for t in transactions[:50]:
        trans_data.append([
            t.get('time_slot', 'N/A'),
            t.get('transaction_type', 'N/A').upper(),
            f"{t.get('quantity_mw', 0):.3f}",
            f"{t.get('rate_per_mwh', 0):.2f}",
            f"₹{t.get('amount', 0):,.2f}"
        ])
    
    trans_table = Table(trans_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1.2*inch, 1.2*inch])
    trans_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(trans_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_daily_trading_excel(transactions: List[Dict], date: str) -> BytesIO:
    """Generate Excel report for daily trading"""
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Header
    ws_summary['A1'] = f"Daily Trading Report - {date}"
    ws_summary['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws_summary['A1'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    ws_summary.merge_cells('A1:E1')
    
    # Summary statistics
    total_volume = sum(t.get('quantity_mw', 0) for t in transactions)
    total_amount = sum(t.get('amount', 0) for t in transactions)
    buy_count = len([t for t in transactions if t.get('transaction_type') == 'buy'])
    sell_count = len([t for t in transactions if t.get('transaction_type') == 'sell'])
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', len(transactions)],
        ['Buy Transactions', buy_count],
        ['Sell Transactions', sell_count],
        ['Total Volume (MW)', round(total_volume, 2)],
        ['Total Amount (₹)', round(total_amount, 2)],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")
    
    # Transactions sheet
    ws_trans = wb.create_sheet("Transactions")
    ws_trans.append(['Date', 'Time Slot', 'Type', 'Report Type', 'Quantity (MW)', 'Rate (₹/MWh)', 'Amount (₹)'])
    
    # Header style
    for cell in ws_trans[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for t in transactions:
        ws_trans.append([
            t.get('date', ''),
            t.get('time_slot', ''),
            t.get('transaction_type', '').upper(),
            t.get('report_type', ''),
            round(t.get('quantity_mw', 0), 3),
            round(t.get('rate_per_mwh', 0), 2),
            round(t.get('amount', 0), 2),
        ])
    
    # Auto-width columns
    for ws in [ws_summary, ws_trans]:
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            for cell in column:
                try:
                    # Skip merged cells
                    if isinstance(cell, openpyxl.cell.cell.MergedCell):
                        continue
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_energy_schedule_pdf(days: List[Dict]) -> BytesIO:
    """Generate PDF report for energy schedule analysis"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#10b981'),
        spaceAfter=30,
        alignment=1
    )
    
    # Title
    title = Paragraph("Energy Schedule Analysis Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Summary
    total_savings = sum(d.get('energy_savings_mwh', 0) for d in days)
    total_cost = sum(d.get('total_cost', 0) for d in days)
    avg_ctu_losses = sum(d.get('ctu_losses_percent', 0) for d in days) / len(days) if days else 0
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Days Calculated', str(len(days))],
        ['Total Energy Savings', f"{total_savings:.2f} MWh"],
        ['Total Cost', f"₹{total_cost:,.2f}"],
        ['Average CTU Losses', f"{avg_ctu_losses:.2f}%"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Daily breakdown
    elements.append(Paragraph("Daily Breakdown", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    daily_data = [['Date', 'Energy Savings (MWh)', 'CTU Losses %', 'Total Cost (₹)']]
    for d in days:
        daily_data.append([
            d.get('trading_date', 'N/A'),
            f"{d.get('energy_savings_mwh', 0):.2f}",
            f"{d.get('ctu_losses_percent', 0):.2f}%",
            f"₹{d.get('total_cost', 0):,.2f}"
        ])
    
    daily_table = Table(daily_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.5*inch])
    daily_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    
    elements.append(daily_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer
