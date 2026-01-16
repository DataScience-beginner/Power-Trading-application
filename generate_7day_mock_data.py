#!/usr/bin/env python3
"""
Generate 7 Days of Complete Mock Trading Data
==============================================

Creates realistic DOR and SCH files for 7 consecutive days with:
- 3 DOR types: GDAM, DAM, RTM (for each day)
- 3 SCH types: GDAM, DAM, RTM (for each day)
- Based on actual Excel template from Data folder
- Ready for upload to Railway PostgreSQL

Total: 7 days × 6 files = 42 files
"""

import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment
import random

# Configuration
START_DATE = datetime(2026, 1, 10)  # Jan 10, 2026
NUM_DAYS = 7
OUTPUT_DIR = Path("mock_data_7days")

# Client configuration (using one main client for simplicity)
CLIENT = {
    "entity_id": "A2AR0NPT0027",
    "entity_name": "Grasim Industries Limited",
    "portfolio_code": "NPT0019_TN0",
    "state_code": "TN",
    "buyer_seller_id": "S1TN0NPT0019"
}

# Template file from Data folder
TEMPLATE_FILE = Path("Data/RTM_IEX130126DOR_NPT0019_TN0_Grasim_Industries_Limited.XLS")


def create_dor_file(trading_date: datetime, market_type: str, output_dir: Path) -> Path:
    """Create DOR (Daily Obligation Report) Excel file"""
    
    # Filename format: [MARKET]_IEX[DDMMYY]DOR_[PORTFOLIO]_[CLIENT].xlsx
    date_str = trading_date.strftime("%d%m%y")
    filename = f"{market_type}_IEX{date_str}DOR_{CLIENT['portfolio_code']}_{CLIENT['entity_name'].replace(' ', '_')}.xlsx"
    output_path = output_dir / filename
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DOR Report"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Header information
    ws['A1'] = "Daily Obligation Report (DOR)"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f"Market Type: {market_type}"
    ws['A3'] = f"Trading Date: {trading_date.strftime('%d-%b-%Y')}"
    ws['A4'] = f"Delivery Date: {(trading_date + timedelta(days=1)).strftime('%d-%b-%Y')}"
    ws['A5'] = f"Entity ID: {CLIENT['entity_id']}"
    ws['A6'] = f"Entity Name: {CLIENT['entity_name']}"
    ws['A7'] = f"Portfolio Code: {CLIENT['portfolio_code']}"
    ws['A8'] = f"Buyer/Seller ID: {CLIENT['buyer_seller_id']}"
    
    # Transaction header (row 10)
    ws['A10'] = "Time Block"
    ws['B10'] = "Buy Qty (MWh)"
    ws['C10'] = "Buy Price (Rs)"
    ws['D10'] = "Sell Qty (MWh)"
    ws['E10'] = "Sell Price (Rs)"
    
    for cell in ['A10', 'B10', 'C10', 'D10', 'E10']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Generate 96 time slots (15-minute intervals)
    row = 11
    for hour in range(24):
        for minute in [0, 15, 30, 45]:
            start_time = f"{hour:02d}:{minute:02d}"
            end_minute = minute + 15
            end_hour = hour
            if end_minute == 60:
                end_minute = 0
                end_hour = (hour + 1) % 24
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            time_block = f"{start_time} - {end_time}"
            
            # Generate realistic trading data
            buy_qty = round(random.uniform(5, 50), 2) if random.random() > 0.3 else 0
            buy_price = round(random.uniform(3.5, 5.5), 2) if buy_qty > 0 else 0
            sell_qty = round(random.uniform(2, 30), 2) if random.random() > 0.4 else 0
            sell_price = round(random.uniform(4.0, 6.0), 2) if sell_qty > 0 else 0
            
            ws[f'A{row}'] = time_block
            ws[f'B{row}'] = buy_qty
            ws[f'C{row}'] = buy_price
            ws[f'D{row}'] = sell_qty
            ws[f'E{row}'] = sell_price
            
            row += 1
    
    # Charges section (after transactions)
    charges_row = row + 2
    ws[f'A{charges_row}'] = "Charges Summary"
    ws[f'A{charges_row}'].font = Font(bold=True)
    
    ws[f'A{charges_row + 1}'] = "NLDC Fee"
    ws[f'B{charges_row + 1}'] = round(random.uniform(1000, 5000), 2)
    
    ws[f'A{charges_row + 2}'] = "CTU Charges"
    ws[f'B{charges_row + 2}'] = round(random.uniform(10000, 50000), 2)
    
    ws[f'A{charges_row + 3}'] = "Other Charges"
    ws[f'B{charges_row + 3}'] = round(random.uniform(500, 2000), 2)
    
    # Save file
    wb.save(output_path)
    print(f"✅ Created DOR: {filename}")
    return output_path


def create_sch_file(trading_date: datetime, market_type: str, output_dir: Path) -> Path:
    """Create SCH (Scheduling Report) Excel file"""
    
    # Filename format: IEX[DDMMYY]SCH_[MARKET]_[PORTFOLIO]_[CLIENT].xlsx
    date_str = trading_date.strftime("%d%m%y")
    
    if market_type == "DAM" and random.random() > 0.5:
        # Some files without market prefix
        filename = f"IEX{date_str}SCH_{CLIENT['portfolio_code']}_{CLIENT['entity_name'].replace(' ', '_')}.xlsx"
    else:
        filename = f"IEX{date_str}SCH_{market_type}_{CLIENT['portfolio_code']}_{CLIENT['entity_name'].replace(' ', '_')}.xlsx"
    
    output_path = output_dir / filename
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scheduling Report"
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    
    # Header information
    ws['A1'] = "Scheduling Report (SCH)"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = f"Market Type: {market_type}"
    ws['A3'] = f"Scheduling Date: {trading_date.strftime('%d-%b-%Y')}"
    ws['A4'] = f"Entity ID: {CLIENT['entity_id']}"
    ws['A5'] = f"Entity Name: {CLIENT['entity_name']}"
    ws['A6'] = f"Portfolio Code: {CLIENT['portfolio_code']}"
    
    # Regional Periphery header (row 8)
    ws['A8'] = "Regional Periphery Scheduling"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Time Block"
    ws['B9'] = "Regional Periphery (MWh)"
    
    for cell in ['A9', 'B9']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Generate 96 time slots for regional periphery
    row = 10
    for hour in range(24):
        for minute in [0, 15, 30, 45]:
            start_time = f"{hour:02d}:{minute:02d}"
            end_minute = minute + 15
            end_hour = hour
            if end_minute == 60:
                end_minute = 0
                end_hour = (hour + 1) % 24
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            time_block = f"{start_time} - {end_time}"
            regional_qty = round(random.uniform(10, 80), 2)
            
            ws[f'A{row}'] = time_block
            ws[f'B{row}'] = regional_qty
            
            row += 1
    
    # Interface Point header
    interface_row = row + 2
    ws[f'A{interface_row}'] = "Interface Point Scheduling"
    ws[f'A{interface_row}'].font = Font(bold=True)
    
    ws[f'A{interface_row + 1}'] = "Time Block"
    ws[f'B{interface_row + 1}'] = "Interface Point (MWh)"
    
    for cell in [f'A{interface_row + 1}', f'B{interface_row + 1}']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Generate 96 time slots for interface point
    row = interface_row + 2
    for hour in range(24):
        for minute in [0, 15, 30, 45]:
            start_time = f"{hour:02d}:{minute:02d}"
            end_minute = minute + 15
            end_hour = hour
            if end_minute == 60:
                end_minute = 0
                end_hour = (hour + 1) % 24
            end_time = f"{end_hour:02d}:{end_minute:02d}"
            
            time_block = f"{start_time} - {end_time}"
            interface_qty = round(random.uniform(5, 40), 2)
            
            ws[f'A{row}'] = time_block
            ws[f'B{row}'] = interface_qty
            
            row += 1
    
    # Save file
    wb.save(output_path)
    print(f"✅ Created SCH: {filename}")
    return output_path


def generate_7day_mock_data():
    """Generate complete 7-day mock dataset"""
    
    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"\n{'='*70}")
    print(f"GENERATING 7-DAY MOCK DATA")
    print(f"{'='*70}\n")
    print(f"Output directory: {OUTPUT_DIR.absolute()}")
    print(f"Date range: {START_DATE.strftime('%Y-%m-%d')} to {(START_DATE + timedelta(days=NUM_DAYS-1)).strftime('%Y-%m-%d')}")
    print(f"Client: {CLIENT['entity_name']}\n")
    
    generated_files = []
    
    # Generate files for each day
    for day_offset in range(NUM_DAYS):
        trading_date = START_DATE + timedelta(days=day_offset)
        
        print(f"\n📅 Day {day_offset + 1}: {trading_date.strftime('%Y-%m-%d (%A)')}")
        print("-" * 70)
        
        # Generate DOR files (3 types)
        for market_type in ["GDAM", "DAM", "RTM"]:
            dor_file = create_dor_file(trading_date, market_type, OUTPUT_DIR)
            generated_files.append(dor_file)
        
        # Generate SCH files (3 types)
        for market_type in ["GDAM", "DAM", "RTM"]:
            sch_file = create_sch_file(trading_date, market_type, OUTPUT_DIR)
            generated_files.append(sch_file)
    
    print(f"\n{'='*70}")
    print(f"✅ GENERATION COMPLETE!")
    print(f"{'='*70}")
    print(f"\nTotal files generated: {len(generated_files)}")
    print(f"   DOR files: {len([f for f in generated_files if 'DOR' in str(f)])}")
    print(f"   SCH files: {len([f for f in generated_files if 'SCH' in str(f)])}")
    print(f"\nFiles saved to: {OUTPUT_DIR.absolute()}")
    
    return generated_files


if __name__ == "__main__":
    files = generate_7day_mock_data()
