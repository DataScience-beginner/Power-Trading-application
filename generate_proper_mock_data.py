"""
Generate PROPER mock data with 5 different clients
Creates real Excel files with modified cell content (not just filenames)
- 5 clients
- 30 days (January 2026)
- 6 files per day (GDAM-DOR, DAM-DOR, RTM-DOR, GDAM-SCH, DAM-SCH, RTM-SCH)
= 900 files total
"""

import openpyxl
from openpyxl import load_workbook
import subprocess
import tempfile
import os
from datetime import datetime, timedelta
from pathlib import Path
import random

# Configuration
NUM_DAYS = 30
START_DATE = datetime(2026, 1, 1)
OUTPUT_DIR = Path("Data/mock_reports_final")

# Template files
DOR_TEMPLATE = "Data/GDAM_IEX120126DOR_NPT0027_KA0_Mellbro_Sugars_Pvt.xls"
SCH_TEMPLATE = "Data/IEX260114SCH_NPT0019_TN0_Grasim_Industries_Limited.xlsx"

# 5 Different Clients
CLIENTS = [
    {
        "entity_id": "A2AR0NPT0001",
        "entity_name": "Tata Power Company Limited",
        "portfolio_code": "S1MH0NPT0001",
        "portfolio_name": "Tata_Power_Mumbai",
        "code": "NPT0001_MH0"
    },
    {
        "entity_id": "A2AR0NPT0002",
        "entity_name": "Adani Power Limited",
        "portfolio_code": "S1GJ0NPT0002",
        "portfolio_name": "Adani_Power_Gujarat",
        "code": "NPT0002_GJ0"
    },
    {
        "entity_id": "A2AR0NPT0003",
        "entity_name": "Tamil Nadu Generation Corporation",
        "portfolio_code": "S1TN0NPT0003",
        "portfolio_name": "TANGEDCO_Chennai",
        "code": "NPT0003_TN0"
    },
    {
        "entity_id": "A2AR0NPT0004",
        "entity_name": "BSES Rajdhani Power Limited",
        "portfolio_code": "S1DL0NPT0004",
        "portfolio_name": "BSES_Delhi",
        "code": "NPT0004_DL0"
    },
    {
        "entity_id": "A2AR0NPT0005",
        "entity_name": "Bangalore Electricity Supply Company",
        "portfolio_code": "S1KA0NPT0005",
        "portfolio_name": "BESCOM_Bangalore",
        "code": "NPT0005_KA0"
    }
]

def convert_xls_to_xlsx(xls_path):
    """Convert .xls to .xlsx using LibreOffice"""
    temp_dir = tempfile.gettempdir()
    result = subprocess.run([
        'soffice', '--headless', '--convert-to', 'xlsx',
        '--outdir', temp_dir, xls_path
    ], capture_output=True, timeout=30)
    
    if result.returncode == 0:
        base_name = os.path.splitext(os.path.basename(xls_path))[0]
        xlsx_path = os.path.join(temp_dir, f"{base_name}.xlsx")
        if os.path.exists(xlsx_path):
            return xlsx_path
    raise Exception("LibreOffice conversion failed")

def modify_dor_file(wb, trading_date, delivery_date, client):
    """Modify DOR Excel file with new dates and client info"""
    ws = wb.active
    
    trading_str = trading_date.strftime("%d.%m.%Y")
    delivery_str = delivery_date.strftime("%d.%m.%Y")
    
    # Update all cells in first 20 rows
    for row_idx in range(1, 21):
        for col_idx in range(1, 30):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                val = str(cell.value)
                
                # Replace dates
                val = val.replace('12.01.2026', trading_str)
                val = val.replace('13.01.2026', delivery_str)
                
                # Replace client info
                val = val.replace('A2AR0NPT0000', client['entity_id'])
                val = val.replace('NEFA Power Trading Private Limited', client['entity_name'])
                val = val.replace('S1KA0NPT0027', client['portfolio_code'])
                val = val.replace('Mellbro_Sugars_Private_Limited', client['portfolio_name'])
                
                cell.value = val
    
    # Randomize quantities slightly
    for row in ws.iter_rows(min_row=20):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0 and cell.value < 10000:
                variation = random.uniform(0.85, 1.15)
                cell.value = round(cell.value * variation, 2)

def modify_sch_file(wb, scheduling_date, client):
    """Modify SCH Excel file with new date and client info"""
    ws = wb.active
    
    sched_str = scheduling_date.strftime("%d.%m.%Y")
    
    # Update all cells in first 30 rows
    for row_idx in range(1, 31):
        for col_idx in range(1, 20):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                val = str(cell.value)
                
                # Replace dates (SCH uses different format)
                val = val.replace('26.01.2014', sched_str)
                val = val.replace('26-01-2014', scheduling_date.strftime("%d-%m-%Y"))
                
                # Replace client info
                val = val.replace('IEX260114SCH', client['entity_id'])
                val = val.replace('NEFA Power Trading Private Limited', client['entity_name'])
                val = val.replace('NPT0019_TN0', client['code'])
                val = val.replace('Grasim_Industries_Limited', client['portfolio_name'])
                
                cell.value = val
    
    # Randomize quantities
    for row in ws.iter_rows(min_row=30):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0 and cell.value < 10000:
                variation = random.uniform(0.85, 1.15)
                cell.value = round(cell.value * variation, 2)

def create_dor_file(template_xlsx, output_path, trading_date, delivery_date, client, market):
    """Create DOR file (GDAM/DAM/RTM)"""
    wb = load_workbook(template_xlsx)
    modify_dor_file(wb, trading_date, delivery_date, client)
    wb.save(output_path)
    wb.close()

def create_sch_file(template_xlsx, output_path, scheduling_date, client, market):
    """Create SCH file (GDAM/DAM/RTM)"""
    wb = load_workbook(template_xlsx)
    modify_sch_file(wb, scheduling_date, client)
    wb.save(output_path)
    wb.close()

def main():
    print("="*80)
    print("FINAL MOCK DATA GENERATOR - 5 Clients × 30 Days × 6 Files")
    print("="*80)
    
    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"✓ Output directory: {OUTPUT_DIR}")
    
    # Convert templates
    print(f"\n📄 Converting template files...")
    try:
        dor_template_xlsx = convert_xls_to_xlsx(DOR_TEMPLATE)
        print(f"✓ DOR template: {dor_template_xlsx}")
    except Exception as e:
        print(f"❌ DOR template conversion failed: {e}")
        return
    
    # SCH is already .xlsx
    sch_template_xlsx = SCH_TEMPLATE
    print(f"✓ SCH template: {sch_template_xlsx}")
    
    total_files = len(CLIENTS) * NUM_DAYS * 6  # 6 files per day (3 DOR + 3 SCH)
    print(f"\n🔧 Generating {total_files} files...")
    print(f"📅 Clients: {len(CLIENTS)}")
    print(f"📅 Days: {NUM_DAYS} (Jan 1-30, 2026)")
    print(f"📄 Files per day: 6 (GDAM-DOR, DAM-DOR, RTM-DOR, GDAM-SCH, DAM-SCH, RTM-SCH)\n")
    
    success_count = 0
    file_count = 0
    
    for client_idx, client in enumerate(CLIENTS, 1):
        print(f"\n{'='*80}")
        print(f"CLIENT {client_idx}/{len(CLIENTS)}: {client['entity_name']}")
        print(f"{'='*80}")
        
        for day_num in range(NUM_DAYS):
            trading_date = START_DATE + timedelta(days=day_num)
            delivery_date = trading_date + timedelta(days=1)
            date_str = trading_date.strftime("%d%m%y")
            
            if day_num % 5 == 0:  # Progress update every 5 days
                print(f"  Day {day_num+1}/{NUM_DAYS}: {trading_date.strftime('%Y-%m-%d')}")
            
            # Create DOR files (GDAM, DAM, RTM)
            for market in ['GDAM', 'DAM', 'RTM']:
                filename = f"{market}_IEX{date_str}DOR_{client['code']}_{client['portfolio_name']}.xlsx"
                output_path = OUTPUT_DIR / filename
                
                try:
                    create_dor_file(dor_template_xlsx, output_path, trading_date, delivery_date, client, market)
                    success_count += 1
                except Exception as e:
                    print(f"    ❌ {filename}: {e}")
                
                file_count += 1
            
            # Create SCH files (GDAM, DAM, RTM)
            for market in ['GDAM', 'DAM', 'RTM']:
                filename = f"IEX{date_str}SCH_{market}_{client['code']}_{client['portfolio_name']}.xlsx"
                output_path = OUTPUT_DIR / filename
                
                try:
                    create_sch_file(sch_template_xlsx, output_path, delivery_date, client, market)
                    success_count += 1
                except Exception as e:
                    print(f"    ❌ {filename}: {e}")
                
                file_count += 1
    
    print(f"\n{'='*80}")
    print(f"✅ GENERATION COMPLETE")
    print(f"{'='*80}")
    print(f"Success: {success_count}/{file_count} files")
    print(f"Location: {OUTPUT_DIR}")
    print(f"\nNEXT STEPS:")
    print(f"1. Upload to Railway: python upload_to_railway_fast.py")
    print(f"2. Energy schedules will auto-calculate for matching DOR+SCH pairs")
    print(f"{'='*80}")

if __name__ == "__main__":
    main()
