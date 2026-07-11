"""
Generate parser-compatible mock DOR and SCH reports.
Creates a full month of 6 files per day:
- DOR-GDAM, DOR-DAM, DOR-RTM
- SCH-GDAM, SCH-DAM, SCH-RTM
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
import os
import random

# Constants
START_DATE = datetime(2026, 1, 1)  # Start from Jan 1
NUM_DAYS = 30  # Generate full month
OUTPUT_DIR = "/workspaces/Power-Trading-application/Data/mock_reports"
CLIENT_CODE = "NPT0027_KA0"
CLIENT_NAME = "Mellbro_Sugars_Pvt"

def create_output_dir():
    """Create output directory for mock reports"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"✓ Created output directory: {OUTPUT_DIR}")

def generate_realistic_consumption():
    """Generate realistic 96 timeslot consumption data (15-min intervals)"""
    consumption = []
    for hour in range(24):
        for quarter in range(4):
            # Base consumption varies by time of day
            if 0 <= hour < 6:  # Night - lower consumption
                base = random.uniform(800, 1200)
            elif 6 <= hour < 9:  # Morning ramp-up
                base = random.uniform(1500, 2000)
            elif 9 <= hour < 18:  # Peak production hours
                base = random.uniform(2200, 2800)
            elif 18 <= hour < 22:  # Evening decline
                base = random.uniform(1800, 2200)
            else:  # Late evening
                base = random.uniform(1000, 1500)
            
            # Add some randomness
            value = base + random.uniform(-100, 100)
            consumption.append(round(value, 2))
    
    return consumption

def create_dor_report(trading_date, market_type="GDAM"):
    """
    Create mock DOR (Daily Obligation Report) Excel file
    Format matches: GDAM_IEX120126DOR_NPT0027_KA0_Mellbro_Sugars_Pvt.xls
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Obligation"
    
    # Formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    
    # Title
    ws['A1'] = f"IEX - Daily Obligation Summary Report ({market_type})"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Report metadata
    ws['A3'] = "Trading Date:"
    ws['B3'] = trading_date.strftime("%d-%b-%Y")
    ws['A4'] = "Delivery Date:"
    ws['B4'] = trading_date.strftime("%d-%b-%Y")
    ws['A5'] = "Entity ID:"
    ws['B5'] = "A2AR0NPT0000"
    ws['A6'] = "Client Code:"
    ws['B6'] = CLIENT_CODE
    ws['A7'] = "Client Name:"
    ws['B7'] = CLIENT_NAME.replace("_", " ")
    ws['A8'] = "Market Type:"
    ws['B8'] = market_type
    
    # Header row. DOR_Parser reads this lightweight mock layout directly.
    row = 10
    headers = ["Time Slot", "Quantity (MW)", "Rate (₹/MWh)", "Amount (₹)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    
    consumption_data = generate_realistic_consumption()
    market_multiplier = {"GDAM": 0.92, "DAM": 1.00, "RTM": 0.35}[market_type]
    rate_base = {"GDAM": 4200, "DAM": 5000, "RTM": 6200}[market_type]
    total_quantity = 0.0
    total_amount = 0.0

    for timeslot_num in range(1, 97):
        row += 1
        minutes_from_start = (timeslot_num - 1) * 15
        start_time = datetime.combine(trading_date, datetime.min.time()) + timedelta(minutes=minutes_from_start)
        end_time = start_time + timedelta(minutes=15)
        quantity_mw = round((consumption_data[timeslot_num - 1] / 1000.0) * market_multiplier, 3)
        rate = round(rate_base + random.uniform(-300, 450), 2)
        amount = round(quantity_mw * rate * 0.25, 2)
        total_quantity += quantity_mw
        total_amount += amount

        data_row = [
            f"{start_time.strftime('%H:%M')} - {end_time.strftime('%H:%M')}",
            quantity_mw,
            rate,
            amount
        ]

        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    row += 1
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=round(total_quantity, 3)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_amount, 2)).font = Font(bold=True)
    
    # Save file
    date_str = trading_date.strftime("%d%m%y")
    filename = f"{market_type}_IEX{date_str}DOR_{CLIENT_CODE}_{CLIENT_NAME}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath

def create_sch_report(trading_date, market_type="GDAM"):
    """
    Create mock SCH (Scheduling) Excel file
    Format matches: IEX260114SCH_NPT0019_TN0_Grasim_Industries_Limited.xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Formatting
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    # Title
    ws['A1'] = f"IEX - Scheduling Report ({market_type})"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Report metadata
    ws['A3'] = "Scheduling Date:"
    ws['B3'] = trading_date.strftime("%d-%b-%Y")
    ws['A4'] = "Client Code:"
    ws['B4'] = CLIENT_CODE
    ws['A5'] = "Client Name:"
    ws['B5'] = CLIENT_NAME.replace("_", " ")
    ws['A6'] = "Region:"
    ws['B6'] = "Southern Region"
    
    # Timeslot header
    row = 8
    headers = ["Timeslot", "Start Time", "End Time", "Scheduled (kWh)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    
    # Generate 96 timeslots (15-minute intervals)
    consumption_data = generate_realistic_consumption()
    
    for timeslot_num in range(1, 97):
        row += 1
        # Calculate time
        minutes_from_start = (timeslot_num - 1) * 15
        start_time = datetime.combine(trading_date, datetime.min.time()) + timedelta(minutes=minutes_from_start)
        end_time = start_time + timedelta(minutes=15)
        
        # Data
        data = [
            timeslot_num,
            start_time.strftime("%H:%M"),
            end_time.strftime("%H:%M"),
            consumption_data[timeslot_num - 1]
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = center
            if col == 4:  # Consumption column
                cell.number_format = '#,##0.00'
    
    # Summary row
    row += 2
    ws.cell(row=row, column=1, value="Total Daily Consumption:").font = Font(bold=True)
    total_cell = ws.cell(row=row, column=4, value=sum(consumption_data))
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    
    # Save file
    date_str = trading_date.strftime("%d%m%y")
    if market_type == "DAM":
        filename = f"IEX{date_str}SCH_{CLIENT_CODE}_{CLIENT_NAME}.xlsx"
    else:
        filename = f"{market_type}_IEX{date_str}SCH_{CLIENT_CODE}_{CLIENT_NAME}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath

def generate_all_mock_reports():
    """Generate complete set of mock reports for testing"""
    create_output_dir()
    
    print(f"\n🔧 Generating {NUM_DAYS} days of mock reports...")
    print(f"📅 Date range: {START_DATE.strftime('%Y-%m-%d')} to {(START_DATE + timedelta(days=NUM_DAYS-1)).strftime('%Y-%m-%d')}\n")
    
    generated_files = []
    
    for day in range(NUM_DAYS):
        current_date = START_DATE + timedelta(days=day)
        print(f"Day {day + 1}/{NUM_DAYS}: {current_date.strftime('%Y-%m-%d')}")
        
        # Generate DOR reports (3 types: GDAM, DAM, RTM)
        for market_type in ["GDAM", "DAM", "RTM"]:
            filepath = create_dor_report(current_date, market_type)
            generated_files.append(filepath)
            print(f"  ✓ Created DOR-{market_type}: {os.path.basename(filepath)}")
        
        # Generate SCH reports (3 types: GDAM, DAM, RTM)
        for market_type in ["GDAM", "DAM", "RTM"]:
            filepath = create_sch_report(current_date, market_type)
            generated_files.append(filepath)
            print(f"  ✓ Created SCH-{market_type}: {os.path.basename(filepath)}")
    
    print(f"\n✅ Generated {len(generated_files)} mock report files")
    print(f"📁 Location: {OUTPUT_DIR}")
    
    return generated_files

if __name__ == "__main__":
    print("=" * 80)
    print("MOCK DATA GENERATOR - Energy Trading Reports")
    print("=" * 80)
    
    files = generate_all_mock_reports()
    
    print("\n" + "=" * 80)
    print("NEXT STEPS:")
    print("=" * 80)
    print("1. Upload these files to the database using: python3 upload_mock_reports.py")
    print("2. Test calculation engine with the uploaded data")
    print("3. Verify energy schedule calculations in dashboard")
    print("=" * 80)
