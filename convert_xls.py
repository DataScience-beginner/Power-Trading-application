import xlrd
import openpyxl
from datetime import datetime
import os

# Patch xlrd formula bug
import xlrd.formula
original = xlrd.formula.evaluate_name_formula
def patched(book, nobj, namex, blah, level=0):
    try:
        return original(book, nobj, namex, blah, level)
    except AssertionError:
        return None
xlrd.formula.evaluate_name_formula = patched

xls_path = 'Data/RTM_IEX130126DOR_NPT0019_TN0_Grasim_Industries_Limited.XLS'
print(f'Opening: {xls_path}')
print(f'File exists: {os.path.exists(xls_path)}')
print(f'File size: {os.path.getsize(xls_path)}')

wb_old = xlrd.open_workbook(xls_path, formatting_info=False)
ws_old = wb_old.sheet_by_index(0)
print(f'Sheet: {ws_old.name}, Rows: {ws_old.nrows}, Cols: {ws_old.ncols}')

wb_new = openpyxl.Workbook()
ws_new = wb_new.active

count = 0
for row_idx in range(ws_old.nrows):
    for col_idx in range(ws_old.ncols):
        cell = ws_old.cell(row_idx, col_idx)
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            continue
        count += 1
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            ws_new.cell(row=row_idx+1, column=col_idx+1, value=cell.value)
        elif cell.ctype == xlrd.XL_CELL_DATE:
            date_tuple = xlrd.xldate_as_tuple(cell.value, wb_old.datemode)
            ws_new.cell(row=row_idx+1, column=col_idx+1, value=datetime(*date_tuple))
        else:
            ws_new.cell(row=row_idx+1, column=col_idx+1, value=str(cell.value) if cell.value else '')

print(f'Copied {count} non-empty cells')

xlsx_path = 'Data/RTM_IEX130126DOR_NPT0019_TN0_Grasim_Industries_Limited.xlsx'
print(f'Saving to: {xlsx_path}')
wb_new.save(xlsx_path)
print(f'File exists: {os.path.exists(xlsx_path)}')
if os.path.exists(xlsx_path):
    print(f'File size: {os.path.getsize(xlsx_path)}')