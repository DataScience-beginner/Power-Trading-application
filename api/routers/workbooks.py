"""Workbook auth, upload, and solar-working routes."""

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database.config import get_db
from api.security.chat_auth import get_current_user
from database.chatbot_models import AppUser
from database.models import WorkbookResultRow, WorkbookUploadRecord
from api.security.upload_security import max_upload_bytes, scan_file, validate_workbook


router = APIRouter(tags=["workbooks"])

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_DIR = PROJECT_ROOT / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

class WorkbookUserProfile(BaseModel):
    """Workbook portal user profile."""

    id: str
    tenant_id: str | None
    email: str
    full_name: str
    role_codes: list[str]


class ParsedSheetSummary(BaseModel):
    """Summary for one parsed workbook sheet."""

    sheet_name: str
    sheet_type: str
    status: str
    row_count: int | None = None
    validation_summary: str | None = None


class WorkbookRowResponse(BaseModel):
    """Calculated workbook row returned to the frontend."""

    reading_date: str
    tneb_total: float
    iex_total: float
    solar_total: float
    tneb_balance: float
    banking_balance: float


class WorkbookUploadApiResponse(BaseModel):
    """Workbook upload and calculation response."""

    workbook_id: str
    file_name: str
    workbook_month: str | None
    status: str
    sheet_summaries: list[ParsedSheetSummary]
    calculation_summary: dict[str, Any]
    preview_rows: list[WorkbookRowResponse]


class WorkbookListItemResponse(BaseModel):
    """Workbook list item."""

    workbook_id: str
    file_name: str
    workbook_month: str | None
    status: str
    uploaded_at: str
    uploaded_by_user_id: str | None
    solar_working_rows: int


class WorkbookResultsApiResponse(BaseModel):
    """Workbook solar-working result rows."""

    workbook_id: str
    workbook_month: str | None
    status: str
    rows: list[WorkbookRowResponse]


def get_workbook_current_user(user: AppUser = Depends(get_current_user)) -> WorkbookUserProfile:
    """Map the enterprise identity into workbook permissions without a second login system."""
    role_codes = ["platform_admin", "tenant_admin"] if user.role == "platform_admin" else ["client_viewer"]
    return WorkbookUserProfile(
        id=user.id,
        tenant_id=str(user.client_id) if user.client_id else None,
        email=user.email,
        full_name=user.display_name,
        role_codes=role_codes,
    )


def require_workbook_roles(*allowed_roles: str):
    """Require the current workbook user to have at least one allowed role."""

    def dependency(current_user: WorkbookUserProfile = Depends(get_workbook_current_user)) -> WorkbookUserProfile:
        if not set(current_user.role_codes).intersection(set(allowed_roles)):
            raise HTTPException(status_code=403, detail="User does not have the required role.")
        return current_user

    return dependency


def _normalize_workbook_month(file_name: str, result_rows: list[WorkbookRowResponse]) -> str | None:
    if result_rows:
        try:
            parsed = datetime.fromisoformat(result_rows[0].reading_date)
            return parsed.strftime("%Y-%m")
        except ValueError:
            pass

    match = re.search(r"([A-Za-z]{3})[' -]?(\d{2})", file_name)
    if not match:
        return None

    month_map = {
        "jan": "01",
        "feb": "02",
        "mar": "03",
        "apr": "04",
        "may": "05",
        "jun": "06",
        "jul": "07",
        "aug": "08",
        "sep": "09",
        "oct": "10",
        "nov": "11",
        "dec": "12",
    }
    month = month_map.get(match.group(1).lower())
    if not month:
        return None
    year = f"20{match.group(2)}"
    return f"{year}-{month}"


def _parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        trimmed = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(trimmed, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_workbook_rows(file_name: str, content: bytes) -> tuple[
    list[ParsedSheetSummary],
    list[WorkbookRowResponse],
    str | None,
]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet_summaries: list[ParsedSheetSummary] = []
    parsed_rows: list[WorkbookRowResponse] = []

    for sheet in workbook.worksheets:
        sheet_row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            reading_date = _parse_date_value(row[0])
            if reading_date is None:
                continue

            numeric_values = [_to_float(value) for value in row[1:] if _to_float(value) is not None]
            if len(numeric_values) < 4:
                continue

            sheet_row_count += 1
            parsed_rows.append(
                WorkbookRowResponse(
                    reading_date=reading_date.isoformat(),
                    tneb_total=numeric_values[0],
                    iex_total=numeric_values[1],
                    solar_total=numeric_values[2],
                    tneb_balance=numeric_values[3],
                    banking_balance=numeric_values[4] if len(numeric_values) > 4 else 0.0,
                )
            )

        sheet_summaries.append(
            ParsedSheetSummary(
                sheet_name=sheet.title,
                sheet_type="worksheet",
                status="parsed" if sheet_row_count else "ignored",
                row_count=sheet_row_count,
                validation_summary=(
                    f"Parsed {sheet_row_count} candidate rows."
                    if sheet_row_count
                    else "No date-plus-numeric rows detected."
                ),
            )
        )

    parsed_rows.sort(key=lambda item: item.reading_date)

    if not parsed_rows:
        today = datetime.utcnow().date()
        parsed_rows = [
            WorkbookRowResponse(
                reading_date=today.isoformat(),
                tneb_total=0.0,
                iex_total=0.0,
                solar_total=0.0,
                tneb_balance=0.0,
                banking_balance=0.0,
            )
        ]

    workbook_month = _normalize_workbook_month(file_name, parsed_rows)
    return sheet_summaries, parsed_rows, workbook_month


@router.get(
    "/api/v1/workbooks",
    response_model=list[WorkbookListItemResponse],
    summary="List workbooks",
    description="Lists uploaded workbook records visible to the current workbook portal user.",
)
async def read_workbooks(
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
) -> list[WorkbookListItemResponse]:
    """Return workbook upload list."""
    uploads = db.query(WorkbookUploadRecord).order_by(WorkbookUploadRecord.uploaded_at.desc()).all()
    results: list[WorkbookListItemResponse] = []
    for upload in uploads:
        row_count = db.query(WorkbookResultRow).filter(WorkbookResultRow.workbook_id == upload.id).count()
        results.append(
            WorkbookListItemResponse(
                workbook_id=upload.id,
                file_name=upload.file_name,
                workbook_month=upload.workbook_month,
                status=upload.status,
                uploaded_at=upload.uploaded_at.isoformat() if upload.uploaded_at else datetime.utcnow().isoformat(),
                uploaded_by_user_id=upload.uploaded_by,
                solar_working_rows=row_count,
            )
        )
    return results


@router.post(
    "/api/v1/workbooks/upload",
    response_model=WorkbookUploadApiResponse,
    summary="Upload workbook",
    description="Uploads a workbook, parses candidate rows, stores results, and returns preview rows.",
)
async def upload_workbook_v1(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
) -> WorkbookUploadApiResponse:
    """Upload and parse a workbook."""
    content = await file.read(max_upload_bytes() + 1)
    validate_workbook(file.filename, content)

    sheet_summaries, parsed_rows, workbook_month = _parse_workbook_rows(file.filename, content)
    workbook_id = str(uuid4())
    workbook_dir = OUTPUT_DIR / "workbook_uploads"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    stored_path = workbook_dir / f"{workbook_id}.xlsx"
    stored_path.write_bytes(content)
    try:
        scan_file(stored_path)
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise

    upload = WorkbookUploadRecord(
        id=workbook_id,
        file_name=file.filename,
        workbook_month=workbook_month,
        status="calculated",
        uploaded_by=current_user.id,
        stored_file_path=str(stored_path),
    )
    db.add(upload)
    db.flush()

    for row in parsed_rows:
        db.add(
            WorkbookResultRow(
                workbook_id=workbook_id,
                reading_date=datetime.fromisoformat(row.reading_date).date(),
                tneb_total=row.tneb_total,
                iex_total=row.iex_total,
                solar_total=row.solar_total,
                tneb_balance=row.tneb_balance,
                banking_balance=row.banking_balance,
            )
        )

    db.commit()

    return WorkbookUploadApiResponse(
        workbook_id=workbook_id,
        file_name=file.filename,
        workbook_month=workbook_month,
        status="calculated",
        sheet_summaries=sheet_summaries,
        calculation_summary={"row_count": len(parsed_rows)},
        preview_rows=parsed_rows[:10],
    )


@router.get(
    "/api/v1/workbooks/{workbook_id}/solar-working",
    response_model=WorkbookResultsApiResponse,
    summary="Get workbook solar-working rows",
    description="Returns calculated solar-working rows for a workbook.",
)
async def read_workbook_results(
    workbook_id: str,
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
) -> WorkbookResultsApiResponse:
    """Return workbook result rows."""
    upload = db.query(WorkbookUploadRecord).filter(WorkbookUploadRecord.id == workbook_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Workbook not found.")

    rows = (
        db.query(WorkbookResultRow)
        .filter(WorkbookResultRow.workbook_id == workbook_id)
        .order_by(WorkbookResultRow.reading_date.asc())
        .all()
    )
    return WorkbookResultsApiResponse(
        workbook_id=upload.id,
        workbook_month=upload.workbook_month,
        status=upload.status,
        rows=[
            WorkbookRowResponse(
                reading_date=row.reading_date.isoformat(),
                tneb_total=row.tneb_total,
                iex_total=row.iex_total,
                solar_total=row.solar_total,
                tneb_balance=row.tneb_balance,
                banking_balance=row.banking_balance,
            )
            for row in rows
        ],
    )
