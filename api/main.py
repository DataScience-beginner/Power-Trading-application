
"""
FastAPI Backend for Power Trading Application
Enterprise-level API with file upload and data retrieval
Version: 1.0.1
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Body, Header
from database.config import get_db
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import json
import os
from pathlib import Path
from datetime import datetime, date
import sys
from io import BytesIO
from uuid import uuid4
import re

# Add project root to path
sys.path.append(str(Path(__file__).parent.parent))



# --- Admin Database Tree Endpoints ---
from sqlalchemy import inspect, text

# ...existing code...

# Place admin endpoints after get_current_admin definition

# ...existing code...

# JWT and Auth imports

# JWT and Auth imports
from jose import JWTError, jwt
from passlib.context import CryptContext
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from datetime import timedelta
from openpyxl import load_workbook

# JWT Config
SECRET_KEY = "your-very-secret-key"  # Change this in production!
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 1 day

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login")

async def get_current_admin(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    if token_data.username != ADMIN_USERNAME:
        raise credentials_exception
    return {"username": token_data.username}

# JWT and Auth imports
from jose import JWTError, jwt
from passlib.context import CryptContext
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from datetime import timedelta

from parsers.DOR_Parser import GDAMTemplateParser
from parsers.SCH_Parser import SCHTemplateParser
from database.config import get_db, init_db
from database import services as db_services
from database.models import WorkbookUploadRecord, WorkbookResultRow
from backend.ai.weather_fetcher import WeatherFetcher
from backend.ai.power_model import PowerForecastModel
from backend.ai.historical_data_fetcher import HistoricalDataFetcher
from backend.ai.eda_module import WeatherEDA

app = FastAPI(
    title="Power Trading Data API",
    description="Enterprise API for parsing and managing power trading data",
    version="1.0.0"
)

# CORS middleware for web access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify actual origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files - check multiple possible locations
frontend_dir = Path(__file__).parent.parent / "frontend"
frontend_react_dist = Path(__file__).parent.parent / "frontend-react" / "dist"

if frontend_react_dist.exists():
    # Production: serve React build
    app.mount("/assets", StaticFiles(directory=str(frontend_react_dist / "assets")), name="assets")
elif frontend_dir.exists() and (frontend_dir / "static").exists():
    # Legacy: serve old frontend static files
    app.mount("/static", StaticFiles(directory=str(frontend_dir / "static")), name="static")

# Data storage
OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# JWT Config
SECRET_KEY = "your-very-secret-key"  # Change this in production!
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24  # 1 day

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/admin/login")

# Admin endpoints: list tables and view table rows (protected by JWT)
@app.get("/api/admin/tables")
async def list_tables(db: Session = Depends(get_db), current_admin: dict = Depends(get_current_admin)):
    inspector = inspect(db.bind)
    return {"tables": inspector.get_table_names()}


@app.get("/api/admin/table/{table_name}")
async def get_table_data(table_name: str, db: Session = Depends(get_db), current_admin: dict = Depends(get_current_admin), limit: int = 100, offset: int = 0):
    # Basic SQL injection protection
    if not table_name.isidentifier():
        raise HTTPException(status_code=400, detail="Invalid table name")
    sql = text(f"SELECT * FROM {table_name} LIMIT :limit OFFSET :offset")
    result = db.execute(sql, {"limit": limit, "offset": offset})
    columns = result.keys()
    raw_rows = result.fetchall()
    # Convert rows to JSON-serializable values (dates, datetimes)
    def serialize_value(v):
        from datetime import datetime, date
        if isinstance(v, datetime) or isinstance(v, date):
            return v.isoformat()
        try:
            # simple types will pass
            json.dumps(v)
            return v
        except Exception:
            return str(v)

    rows = []
    for row in raw_rows:
        d = {}
        for k, v in zip(columns, row):
            d[k] = serialize_value(v)
        rows.append(d)

    return {"columns": list(columns), "rows": rows}



# Demo admin credentials (replace with DB lookup and hashing in production)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"  # Change this in production!

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: str | None = None


class WorkbookLoginRequest(BaseModel):
    email: str
    password: str
    portal: str


class WorkbookUserProfile(BaseModel):
    id: str
    tenant_id: str | None
    email: str
    full_name: str
    role_codes: List[str]


class WorkbookLoginResponse(BaseModel):
    access_token: str
    token_type: str
    user: WorkbookUserProfile


class ParsedSheetSummary(BaseModel):
    sheet_name: str
    sheet_type: str
    status: str
    row_count: int | None = None
    validation_summary: str | None = None


class WorkbookRowResponse(BaseModel):
    reading_date: str
    tneb_total: float
    iex_total: float
    solar_total: float
    tneb_balance: float
    banking_balance: float


class WorkbookUploadApiResponse(BaseModel):
    workbook_id: str
    file_name: str
    workbook_month: str | None
    status: str
    sheet_summaries: List[ParsedSheetSummary]
    calculation_summary: Dict[str, Any]
    preview_rows: List[WorkbookRowResponse]


class WorkbookListItemResponse(BaseModel):
    workbook_id: str
    file_name: str
    workbook_month: str | None
    status: str
    uploaded_at: str
    uploaded_by_user_id: str | None
    solar_working_rows: int


class WorkbookResultsApiResponse(BaseModel):
    workbook_id: str
    workbook_month: str | None
    status: str
    rows: List[WorkbookRowResponse]


WORKBOOK_DEMO_USERS = {
    "admin@demo.local": {
        "password": "Admin123!",
        "full_name": "Platform Admin",
        "role_codes": ["platform_admin", "tenant_admin"],
        "tenant_id": "demo-tenant",
    },
    "tenantadmin@demo.local": {
        "password": "Tenant123!",
        "full_name": "Tenant Admin",
        "role_codes": ["tenant_admin"],
        "tenant_id": "demo-tenant",
    },
    "client@demo.local": {
        "password": "Client123!",
        "full_name": "Client Viewer",
        "role_codes": ["client_viewer"],
        "tenant_id": "demo-tenant",
    },
}


def create_workbook_access_token(user_email: str) -> str:
    expires_at = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    return jwt.encode({"sub": user_email, "exp": expires_at}, SECRET_KEY, algorithm=ALGORITHM)


def get_workbook_current_user(authorization: str | None = Header(default=None)) -> WorkbookUserProfile:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing bearer token.")

    token = authorization.removeprefix("Bearer ").strip()
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
    except JWTError as exc:
        raise HTTPException(status_code=401, detail="Invalid bearer token.") from exc

    if not email or email not in WORKBOOK_DEMO_USERS:
        raise HTTPException(status_code=401, detail="Unknown workbook user.")

    user = WORKBOOK_DEMO_USERS[email]
    return WorkbookUserProfile(
        id=email,
        tenant_id=user["tenant_id"],
        email=email,
        full_name=user["full_name"],
        role_codes=user["role_codes"],
    )


def require_workbook_roles(*allowed_roles: str):
    def dependency(current_user: WorkbookUserProfile = Depends(get_workbook_current_user)) -> WorkbookUserProfile:
        if not set(current_user.role_codes).intersection(set(allowed_roles)):
            raise HTTPException(status_code=403, detail="User does not have the required role.")
        return current_user

    return dependency


def _normalize_workbook_month(file_name: str, result_rows: List[WorkbookRowResponse]) -> str | None:
    if result_rows:
        try:
            parsed = datetime.fromisoformat(result_rows[0].reading_date)
            return parsed.strftime("%Y-%m")
        except ValueError:
            pass

    match = re.search(r"([A-Za-z]{3})[' -]?(\d{2})", file_name)
    if not match:
        return None

    month_map = {
        "jan": "01", "feb": "02", "mar": "03", "apr": "04", "may": "05", "jun": "06",
        "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    month = month_map.get(match.group(1).lower())
    if not month:
        return None
    year = f"20{match.group(2)}"
    return f"{year}-{month}"


def _parse_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        trimmed = value.strip()
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(trimmed, fmt).date()
            except ValueError:
                continue
    return None


def _to_float(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_workbook_rows(file_name: str, content: bytes) -> tuple[List[ParsedSheetSummary], List[WorkbookRowResponse], str | None]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet_summaries: List[ParsedSheetSummary] = []
    parsed_rows: List[WorkbookRowResponse] = []

    for sheet in workbook.worksheets:
        sheet_row_count = 0
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            reading_date = _parse_date_value(row[0])
            if reading_date is None:
                continue

            numeric_values = [_to_float(value) for value in row[1:] if _to_float(value) is not None]
            if len(numeric_values) < 4:
                continue

            sheet_row_count += 1
            tneb_total = numeric_values[0]
            iex_total = numeric_values[1]
            solar_total = numeric_values[2]
            tneb_balance = numeric_values[3]
            banking_balance = numeric_values[4] if len(numeric_values) > 4 else 0.0
            parsed_rows.append(
                WorkbookRowResponse(
                    reading_date=reading_date.isoformat(),
                    tneb_total=tneb_total,
                    iex_total=iex_total,
                    solar_total=solar_total,
                    tneb_balance=tneb_balance,
                    banking_balance=banking_balance,
                )
            )

        sheet_summaries.append(
            ParsedSheetSummary(
                sheet_name=sheet.title,
                sheet_type="worksheet",
                status="parsed" if sheet_row_count else "ignored",
                row_count=sheet_row_count,
                validation_summary=(
                    f"Parsed {sheet_row_count} candidate rows." if sheet_row_count else "No date-plus-numeric rows detected."
                ),
            )
        )

    parsed_rows.sort(key=lambda item: item.reading_date)

    if not parsed_rows:
        today = datetime.utcnow().date()
        parsed_rows = [
            WorkbookRowResponse(
                reading_date=today.isoformat(),
                tneb_total=0.0,
                iex_total=0.0,
                solar_total=0.0,
                tneb_balance=0.0,
                banking_balance=0.0,
            )
        ]

    workbook_month = _normalize_workbook_month(file_name, parsed_rows)
    return sheet_summaries, parsed_rows, workbook_month

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def authenticate_admin(username: str, password: str):
    if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
        return {"username": username}
    return None

def create_access_token(data: dict, expires_delta: timedelta | None = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_admin(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username)
    except JWTError:
        raise credentials_exception
    if token_data.username != ADMIN_USERNAME:
        raise credentials_exception
    return {"username": token_data.username}

# Admin login endpoint
@app.post("/api/admin/login", response_model=Token)
async def admin_login(form_data: OAuth2PasswordRequestForm = Depends()):
    user = authenticate_admin(form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    access_token = create_access_token(
        data={"sub": user["username"]},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return {"access_token": access_token, "token_type": "bearer"}

# Protected endpoint to verify JWT auth
@app.get("/api/admin/me")
async def read_admin_me(current_admin: dict = Depends(get_current_admin)):
    return {"username": current_admin["username"]}


@app.post("/api/v1/auth/login", response_model=WorkbookLoginResponse)
async def workbook_login(payload: WorkbookLoginRequest):
    user = WORKBOOK_DEMO_USERS.get(payload.email)
    if not user or user["password"] != payload.password:
        raise HTTPException(status_code=401, detail="Incorrect email or password")

    profile = WorkbookUserProfile(
        id=payload.email,
        tenant_id=user["tenant_id"],
        email=payload.email,
        full_name=user["full_name"],
        role_codes=user["role_codes"],
    )
    access_token = create_workbook_access_token(payload.email)
    return WorkbookLoginResponse(access_token=access_token, token_type="bearer", user=profile)


@app.get("/api/v1/workbooks", response_model=List[WorkbookListItemResponse])
async def read_workbooks(
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
):
    uploads = db.query(WorkbookUploadRecord).order_by(WorkbookUploadRecord.uploaded_at.desc()).all()
    results: List[WorkbookListItemResponse] = []
    for upload in uploads:
        row_count = db.query(WorkbookResultRow).filter(WorkbookResultRow.workbook_id == upload.id).count()
        results.append(
            WorkbookListItemResponse(
                workbook_id=upload.id,
                file_name=upload.file_name,
                workbook_month=upload.workbook_month,
                status=upload.status,
                uploaded_at=upload.uploaded_at.isoformat() if upload.uploaded_at else datetime.utcnow().isoformat(),
                uploaded_by_user_id=upload.uploaded_by,
                solar_working_rows=row_count,
            )
        )
    return results


@app.post("/api/v1/workbooks/upload", response_model=WorkbookUploadApiResponse)
async def upload_workbook_v1(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx workbooks are supported.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    sheet_summaries, parsed_rows, workbook_month = _parse_workbook_rows(file.filename, content)
    workbook_id = str(uuid4())
    workbook_dir = OUTPUT_DIR / "workbook_uploads"
    workbook_dir.mkdir(parents=True, exist_ok=True)
    stored_path = workbook_dir / f"{workbook_id}.xlsx"
    stored_path.write_bytes(content)

    upload = WorkbookUploadRecord(
        id=workbook_id,
        file_name=file.filename,
        workbook_month=workbook_month,
        status="calculated",
        uploaded_by=current_user.id,
        stored_file_path=str(stored_path),
    )
    db.add(upload)
    db.flush()

    for row in parsed_rows:
        db.add(
            WorkbookResultRow(
                workbook_id=workbook_id,
                reading_date=datetime.fromisoformat(row.reading_date).date(),
                tneb_total=row.tneb_total,
                iex_total=row.iex_total,
                solar_total=row.solar_total,
                tneb_balance=row.tneb_balance,
                banking_balance=row.banking_balance,
            )
        )

    db.commit()

    return WorkbookUploadApiResponse(
        workbook_id=workbook_id,
        file_name=file.filename,
        workbook_month=workbook_month,
        status="calculated",
        sheet_summaries=sheet_summaries,
        calculation_summary={"row_count": len(parsed_rows)},
        preview_rows=parsed_rows[:10],
    )


@app.get("/api/v1/workbooks/{workbook_id}/solar-working", response_model=WorkbookResultsApiResponse)
async def read_workbook_results(
    workbook_id: str,
    db: Session = Depends(get_db),
    current_user: WorkbookUserProfile = Depends(
        require_workbook_roles("platform_admin", "tenant_admin", "client_viewer")
    ),
):
    upload = db.query(WorkbookUploadRecord).filter(WorkbookUploadRecord.id == workbook_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Workbook not found.")

    rows = (
        db.query(WorkbookResultRow)
        .filter(WorkbookResultRow.workbook_id == workbook_id)
        .order_by(WorkbookResultRow.reading_date.asc())
        .all()
    )
    return WorkbookResultsApiResponse(
        workbook_id=upload.id,
        workbook_month=upload.workbook_month,
        status=upload.status,
        rows=[
            WorkbookRowResponse(
                reading_date=row.reading_date.isoformat(),
                tneb_total=row.tneb_total,
                iex_total=row.iex_total,
                solar_total=row.solar_total,
                tneb_balance=row.tneb_balance,
                banking_balance=row.banking_balance,
            )
            for row in rows
        ],
    )

# Initialize database on startup
@app.on_event("startup")
async def startup_event():
    """Initialize database when app starts"""
    print("🗄️  Initializing database...")
    init_db()
    print("✅ Database ready")
    
    if os.getenv("AUTO_LOAD_MOCK_DATA", "false").lower() == "true":
        try:
            from database.config import SessionLocal
            from database.services import get_all_clients
            
            db = SessionLocal()
            try:
                clients = get_all_clients(db)
                if len(clients) == 0:
                    print("📊 Database is empty, loading mock data...")
                    import subprocess
                    subprocess.run([sys.executable, "scripts/data_generation/generate_mock_reports.py"], check=False)
                    subprocess.run([sys.executable, "scripts/ingestion/upload_mock_reports.py"], check=False)
                    subprocess.run([sys.executable, "scripts/energy_schedule/rebuild_energy_schedules.py"], check=False)
                    print("✅ Mock data loaded")
                else:
                    print(f"✅ Database has {len(clients)} clients already")
            finally:
                db.close()
        except Exception as e:
            print(f"⚠️  Mock data load failed: {e}")
            print("   You can upload files manually via the UI")
    else:
        print("ℹ️  AUTO_LOAD_MOCK_DATA is disabled")

@app.get("/")
async def root():
    """Root endpoint - serves the React app"""
    react_index = Path(__file__).parent.parent / "frontend-react" / "dist" / "index.html"
    if react_index.exists():
        return FileResponse(react_index)
    
    # Fallback to old dashboard
    dashboard_file = frontend_dir / "dashboard.html"
    if dashboard_file.exists():
        return FileResponse(dashboard_file)
    
    return {
        "message": "Power Trading Analytics Dashboard",
        "version": "1.0.0",
        "endpoints": {
            "health": "/api/health",
            "upload": "/api/upload",
            "analytics": "/api/analytics/summary",
            "docs": "/docs"
        }
    }

@app.get("/parser")
async def parser_ui():
    """Parser UI - simple upload and view parsed JSON"""
    html_file = frontend_dir / "index.html"
    if html_file.exists():
        return FileResponse(html_file)
    return {"message": "Parser UI not found"}

@app.get("/energy-schedule")
async def energy_schedule_page():
    """Energy Schedule Dashboard - daily CTU losses and cost analysis"""
    energy_schedule_file = frontend_dir / "energy_schedule.html"
    if energy_schedule_file.exists():
        return FileResponse(energy_schedule_file)
    return {"message": "Energy Schedule page not found"}

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "service": "Power Trading API"
    }

@app.post("/api/data/bulk-upload")
async def bulk_upload_data(
    data: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db)
):
    """
    Bulk upload trading data as JSON (bypasses Excel parsing)
    Expects format matching parsed data structure
    """
    try:
        from database import services
        
        entity_id = data.get("entity_id")
        portfolio_code = data.get("portfolio_code")
        trading_date = datetime.fromisoformat(data.get("trading_date")).date()
        report_type = data.get("report_type")
        transactions = data.get("transactions", [])
        
        # Get or create client
        client = services.get_or_create_client(
            db,
            entity_id=entity_id,
            entity_name=data.get("entity_name", "Unknown")
        )
        
        # Get or create portfolio
        portfolio = services.get_or_create_portfolio(
            db,
            client_id=client.id,
            portfolio_code=portfolio_code,
            portfolio_name=data.get("portfolio_name", portfolio_code)
        )
        
        # Create daily file
        daily_file = services.save_daily_file(
            db,
            portfolio_id=portfolio.id,
            trading_date=trading_date,
            report_type=report_type,
            sub_category=data.get("sub_category", "DAM"),
            delivery_date=trading_date,
            filename=f"{report_type}_{trading_date}_{portfolio_code}.json",
            file_size=len(str(data)),
            parsed_data=data
        )
        
        # Save transactions
        saved_count = services.save_transactions(db, daily_file.id, transactions)
        
        return {
            "success": True,
            "message": f"Uploaded {report_type} with {saved_count} transactions",
            "file_id": daily_file.id,
            "client_id": client.id,
            "portfolio_id": portfolio.id
        }
        
    except Exception as e:
        import traceback
        print(f"❌ Bulk upload error: {traceback.format_exc()}")
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/admin/reset-database")
async def reset_database(db: Session = Depends(get_db)):
    """
    ADMIN ENDPOINT: Reset database by dropping all data
    WARNING: This deletes ALL clients, portfolios, files, transactions, and calculations!
    Use this to start fresh when you need to re-upload corrected data.
    """
    from database.models import (
        Client, Portfolio, DailyFile, Transaction,
        EnergyScheduleMonth, EnergyScheduleDay, MonthlyCalculation
    )
    
    try:
        # Count records before deletion
        clients_count = db.query(Client).count()
        portfolios_count = db.query(Portfolio).count()
        files_count = db.query(DailyFile).count()
        transactions_count = db.query(Transaction).count()
        
        # Delete in reverse dependency order
        db.query(Transaction).delete()
        db.query(MonthlyCalculation).delete()
        db.query(EnergyScheduleDay).delete()
        db.query(EnergyScheduleMonth).delete()
        db.query(DailyFile).delete()
        db.query(Portfolio).delete()
        db.query(Client).delete()
        
        db.commit()
        
        return {
            "status": "success",
            "message": "Database reset complete",
            "deleted": {
                "clients": clients_count,
                "portfolios": portfolios_count,
                "daily_files": files_count,
                "transactions": transactions_count
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Database reset failed: {str(e)}"
        )

@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    client_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Upload and parse Excel trading file
    
    Args:
        file: Excel file (.xls or .xlsx)
        client_id: Optional client identifier
        db: Database session (auto-injected)
        
    Returns:
        Parsed trading data in universal schema
    """
    
    # Log incoming file
    print(f"\n{'='*60}")
    print(f"📥 UPLOAD REQUEST: {file.filename}")
    print(f"Content-Type: {file.content_type}")
    print(f"Size: {file.size if hasattr(file, 'size') else 'unknown'}")
    print(f"{'='*60}\n")
    
    # Validate file type (case-insensitive)
    filename_lower = file.filename.lower()
    if not (filename_lower.endswith('.xls') or filename_lower.endswith('.xlsx')):
        print(f"❌ REJECTED: Invalid file extension for {file.filename}")
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Only .xls and .xlsx files are supported."
        )
    
    print(f"✅ File type validation passed: {file.filename}")
    
    try:
        # Save uploaded file temporarily
        temp_file = OUTPUT_DIR / f"temp_{file.filename}"
        print(f"💾 Saving to temp location: {temp_file}")
        
        with open(temp_file, "wb") as f:
            content = await file.read()
            f.write(content)
        
        print(f"✅ File saved successfully ({len(content)} bytes)")
        
        # Detect file type and select appropriate parser
        filename_upper = file.filename.upper()
        if 'SCH' in filename_upper:
            print(f"📋 Detected SCH (Scheduling) report format")
            parser = SCHTemplateParser(client_id=client_id or "default-client")
        else:
            print(f"📋 Detected GDAM/RTM/DOR report format")
            parser = GDAMTemplateParser(client_id=client_id or "default-client")
        
        # Parse the file
        print(f"🔍 Starting parser for: {temp_file.name}")
        parsed_data = parser.parse_excel(str(temp_file))
        
        print(f"✅ Parsing completed successfully")
        print(f"   - Format: {parsed_data['metadata'].get('report_type')}")
        
        # Log transaction counts based on report type
        if 'scheduling_transactions' in parsed_data:
            print(f"   - Scheduling Transactions: {len(parsed_data['scheduling_transactions'])}")
        else:
            print(f"   - Buy Transactions: {len(parsed_data['buy_transactions'])}")
            print(f"   - Sell Transactions: {len(parsed_data['sell_transactions'])}")
        
        # Save parsed data
        output_filename = f"{Path(file.filename).stem}_parsed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        output_file = OUTPUT_DIR / output_filename
        
        print(f"💾 Saving parsed data to: {output_filename}")
        
        with open(output_file, 'w') as f:
            json.dump(parsed_data, f, indent=2)
        
        print(f"✅ Data saved successfully\n")
        
        # ==================== SAVE TO DATABASE ====================
        print(f"\n🗄️  SAVING TO DATABASE...")
        
        metadata = parsed_data['metadata']
        
        # Step 1: Get or create client
        client = db_services.get_or_create_client(
            db,
            entity_id=metadata.get('entity_id', 'UNKNOWN'),
            entity_name=metadata.get('entity_name', 'Unknown Entity')
        )
        
        # Step 2: Get or create portfolio
        portfolio = db_services.get_or_create_portfolio(
            db,
            client_id=client.id,
            portfolio_code=metadata.get('portfolio_code', metadata.get('entity_id', 'DEFAULT')),
            portfolio_name=metadata.get('portfolio_name')
        )
        
        # Step 3: Save daily file (will REPLACE if already exists for same date+type)
        trading_date = datetime.fromisoformat(metadata['trading_date']).date() if metadata.get('trading_date') else date.today()
        delivery_date = datetime.fromisoformat(metadata['delivery_date']).date() if metadata.get('delivery_date') else trading_date
        
        daily_file = db_services.save_daily_file(
            db,
            portfolio_id=portfolio.id,
            trading_date=trading_date,
            delivery_date=delivery_date,
            main_category=metadata.get('main_category', 'DOR'),
            sub_category=metadata.get('sub_category', 'DAM'),
            report_type=metadata.get('report_type', 'DOR-DAM'),
            original_filename=file.filename,
            file_path=str(output_file),
            parsed_data=parsed_data
        )
        
        # Step 4: Save transactions
        transactions_to_save = []
        if 'scheduling_transactions' in parsed_data and parsed_data['scheduling_transactions']:
            transactions_to_save = [{**txn, 'transaction_type': 'scheduling'} for txn in parsed_data['scheduling_transactions']]
        else:
            for txn in parsed_data.get('buy_transactions', []):
                transactions_to_save.append({**txn, 'transaction_type': 'buy'})
            for txn in parsed_data.get('sell_transactions', []):
                transactions_to_save.append({**txn, 'transaction_type': 'sell'})
        
        txn_count = db_services.save_transactions(db, daily_file.id, transactions_to_save)
        
        print(f"✅ DATABASE SAVE COMPLETE:")
        print(f"   - Client: {client.entity_name}")
        print(f"   - Portfolio: {portfolio.portfolio_code}")
        print(f"   - File ID: {daily_file.id}")
        print(f"   - Transactions: {txn_count}")
        
        # ==================== AUTO-BUILD ENERGY SCHEDULE FROM DB DATA ====================
        energy_schedule_result = None
        try:
            from database.energy_schedule_builder import rebuild_energy_schedule_for_day

            print(f"\n⚡ REBUILDING ENERGY SCHEDULE FROM SAVED DATA...")

            energy_schedule_result = rebuild_energy_schedule_for_day(
                db,
                portfolio_id=portfolio.id,
                trading_date=trading_date
            )
            energy_schedule_result["auto_calculated"] = energy_schedule_result["is_complete"]

            print(f"   ✅ Energy schedule day: {energy_schedule_result['daily_entry_id']}")
            print(f"      Complete: {energy_schedule_result['is_complete']}")
            print(f"      Files found: {energy_schedule_result['files_found']}")
            print(f"      Total scheduled: {energy_schedule_result['total_scheduled_mwh']:.2f} MWh")
            print(f"      Total consumption: {energy_schedule_result['total_consumption_after_losses_mwh']:.2f} MWh")
        except Exception as e:
            print(f"\n   ⚠️  Energy Schedule rebuild skipped: {str(e)}")
            energy_schedule_result = {
                "auto_calculated": False,
                "error": str(e)
            }
        
        # Clean up temp file
        temp_file.unlink()
        print(f"\n🗑️  Temp file cleaned up")
        
        return {
            "success": True,
            "message": "File parsed and saved to database successfully",
            "filename": output_filename,
            "database": {
                "client_id": client.id,
                "portfolio_id": portfolio.id,
                "file_id": daily_file.id,
                "transactions_saved": txn_count
            },
            "energy_schedule": energy_schedule_result,
            "data": parsed_data,
            "summary": {
                "trading_date": metadata.get('trading_date'),
                "delivery_date": metadata.get('delivery_date'),
                "entity": metadata.get('entity_name'),
                "portfolio": portfolio.portfolio_code,
                "report_type": metadata.get('report_type'),
                "buy_transactions": parsed_data['summary'].get('total_buy_transactions', 0),
                "sell_transactions": parsed_data['summary'].get('total_sell_transactions', 0),
                "net_amount": parsed_data['summary'].get('net_amount', 0)
            }
        }
        
    except Exception as e:
        # Clean up temp file if it exists
        if 'temp_file' in locals() and temp_file.exists():
            temp_file.unlink()
        
        # Log the full error for debugging
        import traceback
        error_details = traceback.format_exc()
        print(f"\n{'='*60}")
        print(f"❌ ERROR: Failed to parse file {file.filename}")
        print(f"{'='*60}")
        print(error_details)
        print(f"{'='*60}\n")
        
        raise HTTPException(
            status_code=500,
            detail=f"Error parsing file: {str(e)}. Please ensure you're uploading a valid GDAM IEX trading report in .xls or .xlsx format."
        )

@app.get("/api/files")
async def list_files():
    """List all parsed data files"""
    try:
        files = []
        for file_path in OUTPUT_DIR.glob("*_parsed_*.json"):
            stat = file_path.stat()
            files.append({
                "filename": file_path.name,
                "size": stat.st_size,
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
        
        # Sort by modified time, newest first
        files.sort(key=lambda x: x['modified'], reverse=True)
        
        return {
            "success": True,
            "count": len(files),
            "files": files
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error listing files: {str(e)}"
        )

@app.get("/api/data/{filename}")
async def get_data(filename: str):
    """Get parsed data by filename"""
    try:
        file_path = OUTPUT_DIR / filename
        
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {filename}"
            )
        
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        return {
            "success": True,
            "filename": filename,
            "data": data
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )

@app.get("/api/summary/{filename}")
async def get_summary(filename: str):
    """Get summary of parsed data"""
    try:
        file_path = OUTPUT_DIR / filename
        
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {filename}"
            )
        
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        return {
            "success": True,
            "filename": filename,
            "metadata": data.get('metadata', {}),
            "summary": data.get('summary', {}),
            "charges": data.get('charges', {}),
            "transaction_counts": {
                "buy": len(data.get('buy_transactions', [])),
                "sell": len(data.get('sell_transactions', []))
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )

@app.delete("/api/data/{filename}")
async def delete_file(filename: str):
    """Delete a parsed data file"""
    try:
        file_path = OUTPUT_DIR / filename
        
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"File not found: {filename}"
            )
        
        file_path.unlink()
        
        return {
            "success": True,
            "message": f"File deleted: {filename}"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error deleting file: {str(e)}"
        )


# ==================== DATABASE ENDPOINTS ====================

@app.put("/api/transactions/{transaction_id}")
async def update_transaction(
    transaction_id: int,
    updates: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db)
):
    """
    **ADMIN OVERRIDE**: Update transaction values
    
    This allows admins to manually correct any value in a time slot.
    Example: If parser got wrong quantity, admin can fix it here.
    
    Args:
        transaction_id: ID of the transaction to update
        updates: Dictionary of fields to update
            Example: {"quantity_mw": 150.5, "rate_per_mwh": 4250.0}
    """
    try:
        updated_txn = db_services.update_transaction(db, transaction_id, updates)
        
        if not updated_txn:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        return {
            "success": True,
            "message": "Transaction updated successfully",
            "transaction_id": transaction_id,
            "updated_fields": list(updates.keys())
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating transaction: {str(e)}")


@app.get("/api/portfolios/{portfolio_code}/daily-files")
async def get_portfolio_daily_files(
    portfolio_code: str,
    trading_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get all daily files for a portfolio.
    If date provided, returns files for that date (max 6).
    
    This shows which files are uploaded for each day.
    """
    try:
        portfolio = db_services.get_portfolio_by_code(db, portfolio_code)
        
        if not portfolio:
            raise HTTPException(status_code=404, detail="Portfolio not found")
        
        if trading_date:
            # Get files for specific date
            date_obj = datetime.fromisoformat(trading_date).date()
            files = db_services.get_daily_files_by_date(db, portfolio.id, date_obj)
            
            return {
                "success": True,
                "portfolio": portfolio_code,
                "trading_date": trading_date,
                "file_count": len(files),
                "max_files": 6,
                "files": [
                    {
                        "id": f.id,
                        "report_type": f.report_type,
                        "main_category": f.main_category,
                        "sub_category": f.sub_category,
                        "original_filename": f.original_filename,
                        "uploaded_at": f.uploaded_at.isoformat(),
                        "transaction_count": len(f.transactions)
                    }
                    for f in files
                ]
            }
        else:
            # Get all files for portfolio
            return {
                "success": True,
                "portfolio": portfolio_code,
                "message": "Provide trading_date parameter to see files for a specific date"
            }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching files: {str(e)}")


@app.get("/api/files/{file_id}/transactions")
async def get_file_transactions(
    file_id: int,
    db: Session = Depends(get_db)
):
    """
    Get all transactions for a specific file.
    Returns all 96 time slots with data.
    """
    try:
        transactions = db_services.get_transactions_by_file(db, file_id)
        
        return {
            "success": True,
            "file_id": file_id,
            "transaction_count": len(transactions),
            "transactions": [
                {
                    "id": txn.id,
                    "date": txn.date.isoformat(),
                    "time_slot": txn.time_slot,
                    "transaction_type": txn.transaction_type,
                    # Buy fields
                    "quantity_mw": txn.quantity_mw,
                    "rate_per_mwh": txn.rate_per_mwh,
                    "amount": txn.amount,
                    # Sell fields  
                    "solar_quantity_mw": txn.solar_quantity_mw,
                    "non_solar_quantity_mw": txn.non_solar_quantity_mw,
                    "hydro_quantity_mw": txn.hydro_quantity_mw,
                    "total_quantity_mw": txn.total_quantity_mw,
                    # Scheduling fields
                    "regional_injection_mw": txn.regional_injection_mw,
                    "regional_drawal_mw": txn.regional_drawal_mw,
                    "regional_net_mw": txn.regional_net_mw,
                    "interface_injection_mw": txn.interface_injection_mw,
                    "interface_drawal_mw": txn.interface_drawal_mw,
                    "interface_net_mw": txn.interface_net_mw,
                    "total_losses_mw": txn.total_losses_mw
                }
                for txn in transactions
            ]
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching transactions: {str(e)}")


@app.get("/api/clients")
async def get_clients(db: Session = Depends(get_db)):
    """Get all clients"""
    try:
        clients = db_services.get_all_clients(db)
        return {
            "success": True,
            "count": len(clients),
            "clients": [
                {
                    "id": c.id,
                    "entity_id": c.entity_id,
                    "entity_name": c.entity_name,
                    "portfolio_count": len(c.portfolios)
                }
                for c in clients
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching clients: {str(e)}")


@app.get("/api/transactions/all")
async def get_all_transactions(
    portfolio_code: str = None,
    start_date: str = None,
    end_date: str = None,
    report_type: str = None,
    db: Session = Depends(get_db)
):
    """Get all transactions with optional filters"""
    try:
        from database.models import Transaction, DailyFile, Portfolio
        from datetime import datetime
        
        query = db.query(Transaction).join(DailyFile).join(Portfolio)
        
        # Apply filters
        if portfolio_code:
            query = query.filter(Portfolio.portfolio_code == portfolio_code)
        
        if start_date:
            start = datetime.fromisoformat(start_date).date()
            query = query.filter(Transaction.date >= start)
        
        if end_date:
            end = datetime.fromisoformat(end_date).date()
            query = query.filter(Transaction.date <= end)
        
        if report_type:
            query = query.filter(DailyFile.report_type.like(f"{report_type}%"))
        
        # Get total count before limiting
        total_count = query.count()
        
        # Limit to 10000 for performance (configurable via query param later)
        transactions = query.order_by(Transaction.date.desc(), Transaction.time_slot).limit(10000).all()
        
        return {
            "success": True,
            "total_count": total_count,
            "count": len(transactions),
            "transactions": [
                {
                    "id": t.id,
                    "file_id": t.daily_file_id,
                    "date": t.date.isoformat(),
                    "time_slot": t.time_slot,
                    "transaction_type": t.transaction_type,
                    "type": "BUY" if "buy" in t.transaction_type.lower() else "SELL" if "sell" in t.transaction_type.lower() else "SCHEDULE",
                    "report_type": t.daily_file.report_type,
                    "portfolio_code": t.daily_file.portfolio.portfolio_code,
                    "entity_name": t.daily_file.portfolio.client.entity_name,
                    "entity_id": t.daily_file.portfolio.client.entity_id,
                    "quantity": t.quantity_mw,
                    "quantity_mw": t.quantity_mw,
                    "rate": t.rate_per_mwh,
                    "rate_per_mwh": t.rate_per_mwh,
                    "amount": t.amount,
                    "solar_quantity_mw": t.solar_quantity_mw,
                    "non_solar_quantity_mw": t.non_solar_quantity_mw,
                    "total_quantity_mw": t.total_quantity_mw
                }
                for t in transactions
            ]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching transactions: {str(e)}")


@app.get("/api/analytics/summary")
async def get_analytics_summary(
    portfolio_code: str = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    """Get analytics summary for dashboard"""
    try:
        from database.models import DailyFile, Transaction, Portfolio
        from sqlalchemy import func
        from datetime import datetime
        
        # Build base query
        file_query = db.query(DailyFile).join(Portfolio)
        
        if portfolio_code:
            file_query = file_query.filter(Portfolio.portfolio_code == portfolio_code)
        
        if start_date:
            start = datetime.fromisoformat(start_date).date()
            file_query = file_query.filter(DailyFile.trading_date >= start)
        
        if end_date:
            end = datetime.fromisoformat(end_date).date()
            file_query = file_query.filter(DailyFile.trading_date <= end)
        
        files = file_query.all()
        
        # Count DOR vs SCH
        dor_count = sum(1 for f in files if f.report_type.startswith('DOR'))
        sch_count = sum(1 for f in files if f.report_type.startswith('SCH'))
        
        # Get transaction stats
        txn_query = db.query(Transaction).join(DailyFile).join(Portfolio)
        
        if portfolio_code:
            txn_query = txn_query.filter(Portfolio.portfolio_code == portfolio_code)
        
        if start_date:
            txn_query = txn_query.filter(Transaction.date >= datetime.fromisoformat(start_date).date())
        
        if end_date:
            txn_query = txn_query.filter(Transaction.date <= datetime.fromisoformat(end_date).date())
        
        total_transactions = txn_query.count()
        
        # Calculate net amount
        total_amount = db.query(func.sum(Transaction.amount)).join(DailyFile).join(Portfolio)
        if portfolio_code:
            total_amount = total_amount.filter(Portfolio.portfolio_code == portfolio_code)
        if start_date:
            total_amount = total_amount.filter(Transaction.date >= datetime.fromisoformat(start_date).date())
        if end_date:
            total_amount = total_amount.filter(Transaction.date <= datetime.fromisoformat(end_date).date())
        
        net_amount = total_amount.scalar() or 0
        
        # Get hourly distribution - skip for now to avoid SQL compatibility issues
        hourly_data = []
        
        # Buy vs Sell count
        buy_count = txn_query.filter(Transaction.transaction_type == 'buy').count()
        sell_count = txn_query.filter(Transaction.transaction_type == 'sell').count()
        scheduling_count = txn_query.filter(Transaction.transaction_type == 'scheduling').count()
        
        return {
            "success": True,
            "summary": {
                "dor_files": dor_count,
                "sch_files": sch_count,
                "total_files": len(files),
                "total_transactions": total_transactions,
                "net_amount": float(net_amount),
                "buy_transactions": buy_count,
                "sell_transactions": sell_count,
                "scheduling_transactions": scheduling_count
            },
            "hourly_distribution": [
                {
                    "hour": int(h[0]) if h[0] else 0,
                    "avg_quantity": float(h[1]) if h[1] else 0
                }
                for h in hourly_data
            ] if hourly_data else []
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching analytics: {str(e)}")


@app.get("/api/analytics/dor-vs-sch")
async def get_dor_vs_sch_comparison(
    portfolio_code: str = None,
    trading_date: str = None,
    db: Session = Depends(get_db)
):
    """Get DOR vs SCH comparison for same trading date"""
    try:
        from database.models import DailyFile, Transaction, Portfolio
        from datetime import datetime
        from sqlalchemy import func
        
        if not trading_date:
            raise HTTPException(status_code=400, detail="trading_date parameter required")
        
        date_obj = datetime.fromisoformat(trading_date).date()
        
        # Get DOR files for this date
        dor_query = db.query(DailyFile).join(Portfolio).filter(
            DailyFile.trading_date == date_obj,
            DailyFile.report_type.like('DOR%')
        )
        if portfolio_code:
            dor_query = dor_query.filter(Portfolio.portfolio_code == portfolio_code)
        
        # Get SCH files for this date
        sch_query = db.query(DailyFile).join(Portfolio).filter(
            DailyFile.trading_date == date_obj,
            DailyFile.report_type.like('SCH%')
        )
        if portfolio_code:
            sch_query = sch_query.filter(Portfolio.portfolio_code == portfolio_code)
        
        dor_files = dor_query.all()
        sch_files = sch_query.all()
        
        # Get hourly averages for DOR
        dor_hourly = {}
        for file in dor_files:
            transactions = db.query(Transaction).filter(Transaction.daily_file_id == file.id).all()
            for txn in transactions:
                hour = txn.time_slot.split(' - ')[0]
                if hour not in dor_hourly:
                    dor_hourly[hour] = []
                dor_hourly[hour].append(txn.quantity_mw or 0)
        
        # Get hourly averages for SCH
        sch_hourly = {}
        for file in sch_files:
            transactions = db.query(Transaction).filter(Transaction.daily_file_id == file.id).all()
            for txn in transactions:
                hour = txn.time_slot.split(' - ')[0]
                if hour not in sch_hourly:
                    sch_hourly[hour] = []
                sch_hourly[hour].append(txn.quantity_mw or txn.total_quantity_mw or 0)
        
        # Calculate averages
        dor_data = {hour: sum(vals)/len(vals) for hour, vals in dor_hourly.items()}
        sch_data = {hour: sum(vals)/len(vals) for hour, vals in sch_hourly.items()}
        
        # Generate hourly labels
        hours = sorted(set(list(dor_data.keys()) + list(sch_data.keys())))
        
        return {
            "success": True,
            "trading_date": trading_date,
            "dor_files": len(dor_files),
            "sch_files": len(sch_files),
            "comparison": {
                "hours": hours,
                "dor_values": [dor_data.get(h, 0) for h in hours],
                "sch_values": [sch_data.get(h, 0) for h in hours]
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching DOR vs SCH comparison: {str(e)}")


@app.get("/api/energy-schedule/days")
async def get_energy_schedule_days(
    portfolio_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    complete_only: bool = False,
    db: Session = Depends(get_db)
):
    """
    Get energy schedule days with calculations
    
    Returns list of daily energy schedule records with CTU losses, savings, costs
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        
        if complete_only:
            query = query.filter(EnergyScheduleDay.is_complete == 1)
        
        if start_date:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(EnergyScheduleDay.trading_date >= start)
        
        if end_date:
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(EnergyScheduleDay.trading_date <= end)
        
        days = query.order_by(EnergyScheduleDay.trading_date).all()
        
        result = []
        for day in days:
            result.append({
                "id": day.id,
                "trading_date": str(day.trading_date),
                "is_complete": bool(day.is_complete),
                "has_gdam": bool(day.has_gdam_data),
                "has_dam": bool(day.has_dam_data),
                "has_rtm": bool(day.has_rtm_data),
                "has_sch": bool(day.has_sch_data),
                "total_scheduled_mwh": round(day.total_scheduled_mwh, 2) if day.total_scheduled_mwh else 0,
                "ctu_losses_mwh": round(day.ctu_losses_mwh, 2) if day.ctu_losses_mwh else 0,
                "ctu_losses_percent": round(day.ctu_losses_percent, 2) if day.ctu_losses_percent else 0,
                "energy_savings_mwh": round(day.energy_savings_mwh, 2) if day.energy_savings_mwh else 0,
                "total_cost": round(day.total_cost, 2) if day.total_cost else 0,
                "total_consumption_mwh": round(day.total_consumption_after_losses_mwh, 2) if day.total_consumption_after_losses_mwh else 0
            })
        
        return {
            "success": True,
            "count": len(result),
            "days": result
        }
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/api/calculate/energy-schedule")
async def calculate_energy_schedule(
    request_data: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db)
):
    """
    Calculate energy schedule using Excel template
    Uses the existing DOR/SCH parsers and Excel automation
    """
    try:
        from database.energy_schedule_service import calculator
        from datetime import date as dt_date
        
        calculation_date_str = request_data.get('calculation_date')
        print(f"🔍 Calculate request - calculation_date: {calculation_date_str}")
        
        # Parse calculation date
        if calculation_date_str:
            try:
                calc_date = datetime.fromisoformat(calculation_date_str).date()
            except:
                calc_date = datetime.strptime(calculation_date_str.split('T')[0], '%Y-%m-%d').date()
        else:
            calc_date = dt_date.today()
        
        print(f"📅 Using calculation date: {calc_date}")
        
        # Use the full Excel-based calculator
        result = calculator.calculate_energy_schedule(calc_date, db)
        
        return result
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ Error in calculate_energy_schedule: {error_details}")
        
        return {
            "success": False,
            "message": f"Calculation failed: {str(e)}",
            "error": str(e)
        }


@app.get("/api/energy-schedule/status")
async def get_energy_schedule_status(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get energy schedule calculation status for date range
    
    Args:
        start_date: Start date filter
        end_date: End date filter
        
    Returns:
        List of calculated energy schedules
    """
    try:
        from database.models import Base
        from sqlalchemy import Table, MetaData
        
        # Check if table exists (use actual table name: energyscheduleday)
        from database.models import EnergyScheduleDay
        
        # Query energy schedules (using trading_date, not calculation_date)
        query = db.query(EnergyScheduleDay).order_by(EnergyScheduleDay.trading_date.desc()).limit(30)
        schedules_objs = query.all()
        
        schedules = [{
            "id": s.id,
            "month_sheet_id": s.month_sheet_id,
            "trading_date": str(s.trading_date),
            "scheduled_mwh": float(s.gdam_scheduled_quantity_mwh or 0) + float(s.dam_scheduled_quantity_mwh or 0) + float(s.rtm_scheduled_quantity_mwh or 0),
            "consumption_after_losses_mwh": float(s.sch_consumption_after_losses_mwh or 0),
            "gdam_cost": float(s.gdam_cost) if s.gdam_cost else 0.0,
            "dam_cost": float(s.dam_cost) if s.dam_cost else 0.0,
            "rtm_cost": float(s.rtm_cost) if s.rtm_cost else 0.0,
            "ctu_loss_percentage": float(s.ctu_loss_percentage) if s.ctu_loss_percentage else 0.0
        } for s in schedules_objs]
        
        return {
            "success": True,
            "count": len(schedules),
            "schedules": schedules
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching energy schedule status: {str(e)}")


# ==================== ENERGY SCHEDULE ENDPOINTS (Story 2) ====================

@app.get("/api/energy-schedule/months")
async def get_energy_schedule_months(
    portfolio_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Get all energy schedule month sheets
    
    Story 2.4: Energy Savings Summary - List all monthly sheets
    
    Args:
        portfolio_id: Optional portfolio filter
        
    Returns:
        List of month sheets with summary metrics
    """
    try:
        from database.energy_schedule_crud import get_all_month_sheets
        
        month_sheets = get_all_month_sheets(db, portfolio_id)
        
        result = []
        for sheet in month_sheets:
            # Calculate total NLDC fees and CTU charges from dailies
            total_nldc = 0.0
            total_ctu = 0.0
            for day in sheet.daily_entries:
                total_nldc += (day.gdam_nldc_fee or 0) + (day.dam_nldc_fee or 0) + (day.rtm_nldc_fee or 0)
                total_ctu += (day.gdam_ctu_charges or 0) + (day.dam_ctu_charges or 0) + (day.rtm_ctu_charges or 0)
            
            result.append({
                "id": sheet.id,
                "portfolio_id": sheet.portfolio_id,
                "year": sheet.year,
                "month": sheet.month,
                "month_name": sheet.month_name,
                "total_scheduled_mwh": sheet.total_scheduled_mwh,
                "total_consumption_after_losses_mwh": sheet.total_consumption_after_losses_mwh,
                "total_energy_savings": sheet.total_energy_savings_mwh,  # Frontend expects this name
                "total_gdam_cost": sheet.total_gdam_cost,
                "total_dam_cost": sheet.total_dam_cost,
                "total_rtm_cost": sheet.total_rtm_cost,
                "total_nldc_fees": total_nldc,  # Calculated from daily entries
                "total_ctu_charges": total_ctu,  # Calculated from daily entries
                "total_cost": sheet.total_gdam_cost + sheet.total_dam_cost + sheet.total_rtm_cost,  # Add total cost
                "average_ctu_losses": sheet.average_ctu_losses_percent,  # Frontend expects this name
                "total_days_completed": sheet.days_completed,  # Frontend expects this name
                "is_complete": sheet.is_complete,
                "created_at": sheet.created_at.isoformat() if sheet.created_at else None,
                "updated_at": sheet.updated_at.isoformat() if sheet.updated_at else None
            })
        
        return result  # Return array directly for frontend
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching month sheets: {str(e)}")


@app.get("/api/energy-schedule/days")
async def get_energy_schedule_days_by_date(
    portfolio_id: Optional[int] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get daily entries by date range
    
    Story 4.1: Energy Schedule View - Frontend filtering
    
    Args:
        portfolio_id: Optional portfolio filter
        start_date: Start date (YYYY-MM-DD)
        end_date: End date (YYYY-MM-DD)
        
    Returns:
        List of daily entries with calculations
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        from datetime import datetime
        
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        
        if start_date:
            start = datetime.fromisoformat(start_date).date()
            query = query.filter(EnergyScheduleDay.trading_date >= start)
        
        if end_date:
            end = datetime.fromisoformat(end_date).date()
            query = query.filter(EnergyScheduleDay.trading_date <= end)
        
        daily_entries = query.order_by(EnergyScheduleDay.trading_date).all()
        
        result = []
        for entry in daily_entries:
            result.append({
                "id": entry.id,
                "trading_date": entry.trading_date.isoformat(),
                "portfolio_id": entry.month_sheet.portfolio_id,
                "total_scheduled_mwh": float(entry.total_scheduled_mwh or 0),
                "total_consumption_after_losses_mwh": float(entry.total_consumption_after_losses_mwh or 0),
                "ctu_losses_percent": float(entry.ctu_losses_percent or 0),
                "ctu_losses_mwh": float(entry.ctu_losses_mwh or 0),
                "gdam_cost": float(entry.gdam_cost or 0),
                "dam_cost": float(entry.dam_cost or 0),
                "rtm_cost": float(entry.rtm_cost or 0),
                "total_cost": float(entry.total_cost or 0),
                "energy_savings_mwh": float(entry.energy_savings_mwh or 0),
                "total_nldc_fee": float(entry.total_nldc_fee or 0),
                "total_ctu_charges": float(entry.total_ctu_charges or 0),
                "has_gdam_data": bool(entry.has_gdam_data),
                "has_dam_data": bool(entry.has_dam_data),
                "has_rtm_data": bool(entry.has_rtm_data),
                "has_sch_data": bool(entry.has_sch_data),
                "is_complete": bool(entry.is_complete),
                "calculated_at": entry.calculated_at.isoformat() if entry.calculated_at else None
            })
        
        return {
            "success": True,
            "count": len(result),
            "days": result
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching daily entries: {str(e)}")


@app.get("/api/energy-schedule/months/{month_sheet_id}/days")
async def get_energy_schedule_days_by_month(
    month_sheet_id: int,
    db: Session = Depends(get_db)
):
    """
    Get all daily entries for a month sheet
    
    Story 2.1: CTU Losses % - Day-wise breakdown
    
    Args:
        month_sheet_id: Month sheet ID
        
    Returns:
        List of daily entries with calculations
    """
    try:
        from database.energy_schedule_crud import get_all_daily_entries
        
        daily_entries = get_all_daily_entries(db, month_sheet_id)
        
        result = []
        for entry in daily_entries:
            result.append({
                "id": entry.id,
                "trading_date": entry.trading_date.isoformat(),
                "day_of_month": entry.day_of_month,
                # DOR Data
                "gdam": {
                    "nldc_fee": entry.gdam_nldc_fee,
                    "ctu_charges": entry.gdam_ctu_charges,
                    "cost": entry.gdam_cost,
                    "scheduled_mwh": entry.gdam_scheduled_quantity_mwh
                },
                "dam": {
                    "nldc_fee": entry.dam_nldc_fee,
                    "ctu_charges": entry.dam_ctu_charges,
                    "cost": entry.dam_cost,
                    "scheduled_mwh": entry.dam_scheduled_quantity_mwh
                },
                "rtm": {
                    "nldc_fee": entry.rtm_nldc_fee,
                    "ctu_charges": entry.rtm_ctu_charges,
                    "cost": entry.rtm_cost,
                    "scheduled_mwh": entry.rtm_scheduled_quantity_mwh
                },
                # SCH Data
                "consumption_after_losses_mwh": entry.total_consumption_after_losses_mwh,
                "regional_loss_percent": entry.regional_loss_percent,
                "state_loss_percent": entry.state_loss_percent,
                "combined_loss_percent": entry.combined_loss_percent,
                # Calculated Fields
                "total_scheduled_mwh": entry.total_scheduled_mwh,
                "ctu_losses_percent": entry.ctu_losses_percent,
                "ctu_losses_mwh": entry.ctu_losses_mwh,
                "energy_savings_mwh": entry.energy_savings_mwh,
                "total_nldc_fee": entry.total_nldc_fee,
                "total_ctu_charges": entry.total_ctu_charges,
                "total_cost": entry.total_cost,
                # Completeness
                "is_complete": entry.is_complete,
                "has_gdam_data": entry.has_gdam_data,
                "has_dam_data": entry.has_dam_data,
                "has_rtm_data": entry.has_rtm_data,
                "has_sch_data": entry.has_sch_data,
                "calculated_at": entry.calculated_at.isoformat() if entry.calculated_at else None
            })
        
        return {
            "success": True,
            "count": len(result),
            "daily_entries": result
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching daily entries: {str(e)}")


@app.get("/api/energy-schedule/calculations/ctu-losses")
async def get_ctu_losses_analysis(
    portfolio_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Get CTU Losses analysis
    
    Story 2.1: CTU Losses % Calculation
    
    Provides:
    - Day-wise CTU losses %
    - Average CTU losses for period
    - Trend analysis
    
    Args:
        portfolio_id: Optional portfolio filter
        year: Optional year filter
        month: Optional month filter
        
    Returns:
        CTU losses analysis with statistics
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        
        # Build query
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        if year:
            query = query.filter(EnergyScheduleMonth.year == year)
        if month:
            query = query.filter(EnergyScheduleMonth.month == month)
        
        entries = query.filter(EnergyScheduleDay.is_complete == 1).order_by(EnergyScheduleDay.trading_date).all()
        
        # Calculate statistics
        if entries:
            total_scheduled = sum(e.total_scheduled_mwh for e in entries)
            total_losses = sum(e.ctu_losses_mwh for e in entries)
            avg_ctu_losses_percent = (total_losses / total_scheduled * 100) if total_scheduled > 0 else 0
            
            daily_breakdown = []
            for entry in entries:
                daily_breakdown.append({
                    "date": entry.trading_date.isoformat(),
                    "scheduled_mwh": entry.total_scheduled_mwh,
                    "consumption_mwh": entry.total_consumption_after_losses_mwh,
                    "ctu_losses_mwh": entry.ctu_losses_mwh,
                    "ctu_losses_percent": entry.ctu_losses_percent
                })
            
            return {
                "success": True,
                "summary": {
                    "total_scheduled_mwh": total_scheduled,
                    "total_losses_mwh": total_losses,
                    "average_ctu_losses_percent": avg_ctu_losses_percent,
                    "days_analyzed": len(entries)
                },
                "daily_breakdown": daily_breakdown
            }
        else:
            return {
                "success": True,
                "summary": {
                    "total_scheduled_mwh": 0,
                    "total_losses_mwh": 0,
                    "average_ctu_losses_percent": 0,
                    "days_analyzed": 0
                },
                "daily_breakdown": [],
                "message": "No complete entries found for the specified period"
            }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating CTU losses: {str(e)}")


@app.get("/api/energy-schedule/calculations/ctu-charges")
async def get_ctu_charges_summary(
    portfolio_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Get CTU Charges summary
    
    Story 2.2: CTU Charges Calculation
    
    Aggregates CTU transmission charges across markets:
    - GDAM CTU charges
    - DAM CTU charges
    - RTM CTU charges
    - Total CTU charges
    
    Args:
        portfolio_id: Optional portfolio filter
        year: Optional year filter
        month: Optional month filter
        
    Returns:
        CTU charges breakdown by market
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        
        # Build query
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        if year:
            query = query.filter(EnergyScheduleMonth.year == year)
        if month:
            query = query.filter(EnergyScheduleMonth.month == month)
        
        entries = query.filter(EnergyScheduleDay.is_complete == 1).order_by(EnergyScheduleDay.trading_date).all()
        
        # Calculate aggregates
        if entries:
            total_gdam_ctu = sum(e.gdam_ctu_charges for e in entries)
            total_dam_ctu = sum(e.dam_ctu_charges for e in entries)
            total_rtm_ctu = sum(e.rtm_ctu_charges for e in entries)
            total_ctu = sum(e.total_ctu_charges for e in entries)
            
            daily_breakdown = []
            for entry in entries:
                daily_breakdown.append({
                    "date": entry.trading_date.isoformat(),
                    "gdam_ctu_charges": entry.gdam_ctu_charges,
                    "dam_ctu_charges": entry.dam_ctu_charges,
                    "rtm_ctu_charges": entry.rtm_ctu_charges,
                    "total_ctu_charges": entry.total_ctu_charges
                })
            
            return {
                "success": True,
                "summary": {
                    "total_gdam_ctu_charges": total_gdam_ctu,
                    "total_dam_ctu_charges": total_dam_ctu,
                    "total_rtm_ctu_charges": total_rtm_ctu,
                    "total_ctu_charges": total_ctu,
                    "days_analyzed": len(entries)
                },
                "daily_breakdown": daily_breakdown
            }
        else:
            return {
                "success": True,
                "summary": {
                    "total_gdam_ctu_charges": 0,
                    "total_dam_ctu_charges": 0,
                    "total_rtm_ctu_charges": 0,
                    "total_ctu_charges": 0,
                    "days_analyzed": 0
                },
                "daily_breakdown": [],
                "message": "No complete entries found for the specified period"
            }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating CTU charges: {str(e)}")


@app.get("/api/energy-schedule/calculations/nldc-fees")
async def get_nldc_fees_summary(
    portfolio_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Get NLDC Fees aggregation
    
    Story 2.3: NLDC Fee Aggregation
    
    Aggregates NLDC application fees across markets:
    - GDAM NLDC fees
    - DAM NLDC fees
    - RTM NLDC fees
    - Total NLDC fees
    
    Args:
        portfolio_id: Optional portfolio filter
        year: Optional year filter
        month: Optional month filter
        
    Returns:
        NLDC fees breakdown by market
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        
        # Build query
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        if year:
            query = query.filter(EnergyScheduleMonth.year == year)
        if month:
            query = query.filter(EnergyScheduleMonth.month == month)
        
        entries = query.filter(EnergyScheduleDay.is_complete == 1).order_by(EnergyScheduleDay.trading_date).all()
        
        # Calculate aggregates
        if entries:
            total_gdam_nldc = sum(e.gdam_nldc_fee for e in entries)
            total_dam_nldc = sum(e.dam_nldc_fee for e in entries)
            total_rtm_nldc = sum(e.rtm_nldc_fee for e in entries)
            total_nldc = sum(e.total_nldc_fee for e in entries)
            
            daily_breakdown = []
            for entry in entries:
                daily_breakdown.append({
                    "date": entry.trading_date.isoformat(),
                    "gdam_nldc_fee": entry.gdam_nldc_fee,
                    "dam_nldc_fee": entry.dam_nldc_fee,
                    "rtm_nldc_fee": entry.rtm_nldc_fee,
                    "total_nldc_fee": entry.total_nldc_fee
                })
            
            return {
                "success": True,
                "summary": {
                    "total_gdam_nldc_fees": total_gdam_nldc,
                    "total_dam_nldc_fees": total_dam_nldc,
                    "total_rtm_nldc_fees": total_rtm_nldc,
                    "total_nldc_fees": total_nldc,
                    "days_analyzed": len(entries)
                },
                "daily_breakdown": daily_breakdown
            }
        else:
            return {
                "success": True,
                "summary": {
                    "total_gdam_nldc_fees": 0,
                    "total_dam_nldc_fees": 0,
                    "total_rtm_nldc_fees": 0,
                    "total_nldc_fees": 0,
                    "days_analyzed": 0
                },
                "daily_breakdown": [],
                "message": "No complete entries found for the specified period"
            }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating NLDC fees: {str(e)}")


@app.get("/api/energy-schedule/calculations/energy-savings")
async def get_energy_savings_summary(
    portfolio_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """
    Get Energy Savings summary
    
    Story 2.4: Energy Savings Summary
    
    Provides comprehensive energy savings analysis:
    - Total energy savings (MWh)
    - Day-wise breakdown
    - Cost savings analysis
    - Trend data
    
    Args:
        portfolio_id: Optional portfolio filter
        year: Optional year filter
        month: Optional month filter
        
    Returns:
        Energy savings summary with detailed breakdown
    """
    try:
        from database.models import EnergyScheduleDay, EnergyScheduleMonth
        
        # Build query
        query = db.query(EnergyScheduleDay).join(EnergyScheduleMonth)
        
        if portfolio_id:
            query = query.filter(EnergyScheduleMonth.portfolio_id == portfolio_id)
        if year:
            query = query.filter(EnergyScheduleMonth.year == year)
        if month:
            query = query.filter(EnergyScheduleMonth.month == month)
        
        entries = query.filter(EnergyScheduleDay.is_complete == 1).order_by(EnergyScheduleDay.trading_date).all()
        
        # Calculate comprehensive metrics
        if entries:
            total_energy_savings = sum(e.energy_savings_mwh for e in entries)
            total_scheduled = sum(e.total_scheduled_mwh for e in entries)
            total_consumption = sum(e.total_consumption_after_losses_mwh for e in entries)
            total_cost = sum(e.total_cost for e in entries)
            
            daily_breakdown = []
            for entry in entries:
                daily_breakdown.append({
                    "date": entry.trading_date.isoformat(),
                    "scheduled_mwh": entry.total_scheduled_mwh,
                    "consumption_mwh": entry.total_consumption_after_losses_mwh,
                    "energy_savings_mwh": entry.energy_savings_mwh,
                    "ctu_losses_percent": entry.ctu_losses_percent,
                    "total_cost": entry.total_cost
                })
            
            return {
                "success": True,
                "summary": {
                    "total_energy_savings_mwh": total_energy_savings,
                    "total_scheduled_mwh": total_scheduled,
                    "total_consumption_mwh": total_consumption,
                    "total_cost": total_cost,
                    "average_savings_percent": (total_energy_savings / total_scheduled * 100) if total_scheduled > 0 else 0,
                    "days_analyzed": len(entries)
                },
                "daily_breakdown": daily_breakdown,
                "trend": {
                    "min_savings": min(e.energy_savings_mwh for e in entries),
                    "max_savings": max(e.energy_savings_mwh for e in entries),
                    "avg_savings": total_energy_savings / len(entries)
                }
            }
        else:
            return {
                "success": True,
                "summary": {
                    "total_energy_savings_mwh": 0,
                    "total_scheduled_mwh": 0,
                    "total_consumption_mwh": 0,
                    "total_cost": 0,
                    "average_savings_percent": 0,
                    "days_analyzed": 0
                },
                "daily_breakdown": [],
                "message": "No complete entries found for the specified period"
            }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error calculating energy savings: {str(e)}")


# ==================== REPORT GENERATION ENDPOINTS ====================

@app.get("/api/reports/daily-trading/pdf")
async def download_daily_trading_pdf(
    date: str = None,
    portfolio_code: str = None,
    db: Session = Depends(get_db)
):
    """Generate and download PDF report for daily trading"""
    try:
        from database.models import Transaction, DailyFile, Portfolio
        from api.report_generator import generate_daily_trading_pdf
        
        query = db.query(Transaction).join(DailyFile).join(Portfolio)
        
        if date:
            query = query.filter(Transaction.date == datetime.fromisoformat(date).date())
        if portfolio_code:
            query = query.filter(Portfolio.portfolio_code == portfolio_code)
        
        transactions = query.all()
        
        trans_data = [
            {
                "date": t.date.isoformat(),
                "time_slot": t.time_slot,
                "transaction_type": t.transaction_type,
                "report_type": t.daily_file.report_type,
                "quantity_mw": t.quantity_mw,
                "rate_per_mwh": t.rate_per_mwh,
                "amount": t.amount
            }
            for t in transactions
        ]
        
        pdf_buffer = generate_daily_trading_pdf(trans_data, date or datetime.now().strftime("%Y-%m-%d"))
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=Daily_Trading_Report_{date or 'latest'}.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


@app.get("/api/reports/daily-trading/excel")
async def download_daily_trading_excel(
    date: str = None,
    portfolio_code: str = None,
    db: Session = Depends(get_db)
):
    """Generate and download Excel report for daily trading"""
    try:
        from database.models import Transaction, DailyFile, Portfolio
        from api.report_generator import generate_daily_trading_excel
        
        query = db.query(Transaction).join(DailyFile).join(Portfolio)
        
        if date:
            query = query.filter(Transaction.date == datetime.fromisoformat(date).date())
        if portfolio_code:
            query = query.filter(Portfolio.portfolio_code == portfolio_code)
        
        transactions = query.all()
        
        trans_data = [
            {
                "date": t.date.isoformat(),
                "time_slot": t.time_slot,
                "transaction_type": t.transaction_type,
                "report_type": t.daily_file.report_type,
                "quantity_mw": t.quantity_mw,
                "rate_per_mwh": t.rate_per_mwh,
                "amount": t.amount
            }
            for t in transactions
        ]
        
        excel_buffer = generate_daily_trading_excel(trans_data, date or datetime.now().strftime("%Y-%m-%d"))
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Daily_Trading_Report_{date or 'latest'}.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel: {str(e)}")


@app.get("/api/reports/energy-schedule/pdf")
async def download_energy_schedule_pdf(
    portfolio_id: int = None,
    start_date: str = None,
    end_date: str = None,
    db: Session = Depends(get_db)
):
    """Generate and download PDF report for energy schedule analysis"""
    try:
        from database.models import EnergyScheduleDay
        from api.report_generator import generate_energy_schedule_pdf
        
        # Query daily entries directly
        query = db.query(EnergyScheduleDay)
        
        # Filter by portfolio_id if provided
        if portfolio_id:
            from database.models import EnergyScheduleMonth
            query = query.join(EnergyScheduleMonth).filter(
                EnergyScheduleMonth.portfolio_id == portfolio_id
            )
        
        # Filter by date range
        if start_date:
            query = query.filter(EnergyScheduleDay.trading_date >= start_date)
        if end_date:
            query = query.filter(EnergyScheduleDay.trading_date <= end_date)
        
        days = query.order_by(EnergyScheduleDay.trading_date).all()
        
        days_data = [
            {
                "trading_date": day.trading_date.isoformat(),
                "energy_savings_mwh": day.energy_savings_mwh or 0,
                "ctu_losses_percent": day.ctu_losses_percent or 0,
                "total_cost": day.total_cost or 0
            }
            for day in days
        ]
        
        pdf_buffer = generate_energy_schedule_pdf(days_data)
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=Energy_Schedule_Report.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating PDF: {str(e)}")


# ===========================
# AI POWER FORECASTING ENDPOINTS
# ===========================

weather_fetcher = WeatherFetcher()
power_model = PowerForecastModel()
historical_fetcher = HistoricalDataFetcher()
eda_analyzer = WeatherEDA()

@app.get("/api/ai/weather/{client_id}")
async def get_weather_forecast(
    client_id: str,
    lat: float,
    lon: float,
    days_ahead: int = 7
):
    """
    Get weather data for AI power forecasting
    
    Args:
        client_id: Client identifier
        lat: Latitude (e.g., 12.97 for Chennai)
        lon: Longitude (e.g., 80.22 for Chennai)
        days_ahead: Number of forecast days (default 7)
        
    Returns:
        Weather data with historical summary and daily forecast
    """
    try:
        weather_data = weather_fetcher.get_weather_data(lat, lon, days_ahead)
        
        return JSONResponse(content={
            "success": True,
            "client_id": client_id,
            "data": weather_data
        })
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error fetching weather data: {str(e)}"
        )


@app.post("/api/ai/forecast-power/{client_id}")
async def forecast_power_generation(
    client_id: str,
    lat: float = Body(...),
    lon: float = Body(...),
    capacity_kw: float = Body(5000),
    farm_type: str = Body("solar"),
    days_ahead: int = Body(7)
):
    """
    AI-powered power generation forecast
    
    Args:
        client_id: Client identifier
        lat: Latitude
        lon: Longitude
        capacity_kw: Farm capacity in kW (default 5000)
        farm_type: "solar" or "wind" (default "solar")
        days_ahead: Forecast days (default 7)
        
    Returns:
        Power forecast with p10/p50/p90 confidence intervals and recommended bid
    """
    try:
        # Get weather data
        weather_data = weather_fetcher.get_weather_data(lat, lon, days_ahead)
        daily_weather = weather_data.get("daily_forecast", [])
        
        # Run power forecast model
        forecast = power_model.forecast_power(
            daily_weather=daily_weather,
            capacity_kw=capacity_kw,
            farm_type=farm_type
        )
        
        return JSONResponse(content={
            "success": True,
            "client_id": client_id,
            "forecast": forecast,
            "weather_summary": weather_data.get("historical_summary"),
            "location": {"lat": lat, "lon": lon}
        })
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error generating power forecast: {str(e)}"
        )


@app.get("/api/ai/inspect-weather/{client_id}")
async def inspect_weather_data(
    client_id: str,
    lat: float,
    lon: float,
    days_ahead: int = 7
):
    """
    EDA: Inspect weather data before forecasting
    Returns detailed weather stats and data quality checks
    """
    try:
        weather_data = weather_fetcher.get_weather_data(lat, lon, days_ahead)
        historical = weather_data.get("historical_summary", {})
        daily = weather_data.get("daily_forecast", [])
        
        # Calculate statistics
        ghi_values = [d.get("ghi_wh_m2", 0) for d in daily]
        temp_values = [d.get("temp_c", 0) for d in daily]
        precip_values = [d.get("precip_mm", 0) for d in daily]
        cloud_risks = [d.get("cloud_risk", 0) for d in daily]
        
        stats = {
            "ghi_stats": {
                "min": round(min(ghi_values) if ghi_values else 0, 2),
                "max": round(max(ghi_values) if ghi_values else 0, 2),
                "avg": round(sum(ghi_values) / len(ghi_values) if ghi_values else 0, 2)
            },
            "temp_stats": {
                "min": round(min(temp_values) if temp_values else 0, 1),
                "max": round(max(temp_values) if temp_values else 0, 1),
                "avg": round(sum(temp_values) / len(temp_values) if temp_values else 0, 1)
            },
            "precip_stats": {
                "total": round(sum(precip_values), 1),
                "rainy_days": sum(1 for p in precip_values if p > 1)
            },
            "cloud_risk_stats": {
                "avg": round(sum(cloud_risks) / len(cloud_risks) if cloud_risks else 0, 2),
                "high_risk_days": sum(1 for c in cloud_risks if c > 0.6)
            }
        }
        
        quality_checks = {
            "all_dates_present": len(daily) == days_ahead,
            "data_completeness": f"{len(daily)}/{days_ahead} days"
        }
        
        return JSONResponse(content={
            "success": True,
            "client_id": client_id,
            "location": {"lat": lat, "lon": lon},
            "historical_summary": historical,
            "statistics": stats,
            "quality_checks": quality_checks,
            "daily_forecast": daily
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.get("/api/ai/historical-data/{client_id}")
async def get_historical_data(
    client_id: str,
    lat: float,
    lon: float,
    years: int = 10
):
    """
    Step 1: Fetch 10 years of historical weather data with progress logs
    Shows raw data in table format + caches for consistency
    """
    try:
        result = historical_fetcher.get_historical_data(lat, lon, years)
        return JSONResponse(content={
            "success": True,
            "client_id": client_id,
            **result
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@app.post("/api/ai/eda-analysis/{client_id}")
async def perform_eda(
    client_id: str,
    raw_data: List[Dict] = Body(...)
):
    """
    Step 2: Perform EDA on historical data to find patterns
    Analyzes monthly/seasonal trends, correlations, insights
    """
    try:
        analysis = eda_analyzer.analyze(raw_data)
        return JSONResponse(content={
            "success": True,
            "client_id": client_id,
            "eda_results": analysis
        })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
