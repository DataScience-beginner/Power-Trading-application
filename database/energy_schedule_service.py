"""
Energy Schedule Calculation Service
Handles validation, Excel automation, and calculation workflow
"""

from pathlib import Path
from datetime import datetime, timedelta, date
from typing import Dict, List, Optional, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
import shutil
import openpyxl
from openpyxl import load_workbook

from database.models import DailyFile, Transaction, EnergyScheduleDaily, EnergyScheduleHourly
from database.config import get_db

class EnergyScheduleCalculator:
    """Manages energy schedule calculations and Excel automation"""
    
    CALCULATIONS_DIR = Path("calculations")
    TEMPLATES_DIR = Path("templates")
    MASTER_TEMPLATE = TEMPLATES_DIR / "master_energy_schedule.xlsx"
    
    def __init__(self):
        """Initialize calculator"""
        self.CALCULATIONS_DIR.mkdir(exist_ok=True)
        self.TEMPLATES_DIR.mkdir(exist_ok=True)
    
    def validate_files_for_calculation(
        self, 
        calculation_date: date,
        db: Session
    ) -> Dict:
        """
        Validate if required DOR and SCH files exist for calculation
        
        Args:
            calculation_date: The date for calculation (SCH date)
            db: Database session
            
        Returns:
            Dictionary with validation status and file information
        """
        # Calculate dates
        dor_date = calculation_date - timedelta(days=1)  # DOR is 1 day before
        sch_date = calculation_date
        
        # Query DOR files (previous day)
        dor_files = db.query(DailyFile).filter(
            and_(
                DailyFile.trading_date == dor_date,
                DailyFile.report_type.like('DOR%')
            )
        ).all()
        
        # Query SCH files (current day)
        sch_files = db.query(DailyFile).filter(
            and_(
                DailyFile.trading_date == sch_date,
                DailyFile.report_type.like('SCH%')
            )
        ).all()
        
        # Extract market types
        dor_markets = sorted(set([f.market_type for f in dor_files]))
        sch_markets = sorted(set([f.market_type for f in sch_files]))
        
        # Count matching files
        dor_count = len(dor_files)
        sch_count = len(sch_files)
        
        # Determine status
        if dor_count == 0 and sch_count == 0:
            status = "missing_all"
            message = f"No files found. Please upload DOR ({dor_date}) and SCH ({sch_date})"
        elif dor_count == 0:
            status = "missing_dor"
            message = f"Missing DOR files for {dor_date}. Found {sch_count} SCH file(s)"
        elif sch_count == 0:
            status = "missing_sch"
            message = f"Missing SCH files for {sch_date}. Found {dor_count} DOR file(s)"
        else:
            status = "ready"
            message = f"Ready to calculate! Found {dor_count} DOR + {sch_count} SCH files"
        
        return {
            "status": status,
            "message": message,
            "calculation_date": str(calculation_date),
            "dor_date": str(dor_date),
            "sch_date": str(sch_date),
            "files_found": {
                "dor_count": dor_count,
                "sch_count": sch_count,
                "dor_markets": dor_markets,
                "sch_markets": sch_markets,
                "count_summary": f"{dor_count}+{sch_count}"
            },
            "dor_files": [
                {
                    "id": f.id,
                    "market_type": f.market_type,
                    "report_type": f.report_type,
                    "filename": f.original_filename
                }
                for f in dor_files
            ],
            "sch_files": [
                {
                    "id": f.id,
                    "market_type": f.market_type,
                    "report_type": f.report_type,
                    "filename": f.original_filename
                }
                for f in sch_files
            ]
        }
    
    def get_or_create_monthly_excel(self, calculation_date: date) -> Path:
        """
        Get existing monthly Excel file or create from template
        
        Args:
            calculation_date: Date for calculation
            
        Returns:
            Path to monthly Excel file
        """
        year = calculation_date.strftime("%Y")
        month_year = calculation_date.strftime("%b_%Y").upper()  # JAN_2026
        
        # Create year folder if not exists
        year_folder = self.CALCULATIONS_DIR / year
        year_folder.mkdir(exist_ok=True)
        
        # Monthly Excel path
        monthly_file = year_folder / f"{month_year}.xlsx"
        
        # If file doesn't exist, copy from template
        if not monthly_file.exists():
            if not self.MASTER_TEMPLATE.exists():
                raise FileNotFoundError(
                    f"Master template not found at {self.MASTER_TEMPLATE}"
                )
            
            shutil.copy(self.MASTER_TEMPLATE, monthly_file)
            print(f"✅ Created monthly Excel: {monthly_file}")
        else:
            print(f"📊 Using existing Excel: {monthly_file}")
        
        return monthly_file
    
    def get_dor_summary_data(
        self, 
        dor_files: List[DailyFile],
        db: Session
    ) -> Dict:
        """
        Extract summary data from DOR files for Excel rows 4, 5, 6
        
        Args:
            dor_files: List of DOR file records
            db: Database session
            
        Returns:
            Dictionary with GDAM, DAM, RTM summary data
        """
        summary = {
            "GDAM": {"losses_pct": 0, "charges": 0, "cost": 0},
            "DAM": {"losses_pct": 0, "charges": 0, "cost": 0},
            "RTM": {"losses_pct": 0, "charges": 0, "cost": 0}
        }
        
        for dor_file in dor_files:
            market = dor_file.market_type
            
            # Get transactions for this file
            transactions = db.query(Transaction).filter(
                Transaction.daily_file_id == dor_file.id
            ).all()
            
            if not transactions:
                continue
            
            # Calculate totals
            total_quantity = sum(t.quantity_mw for t in transactions if t.quantity_mw)
            total_amount = sum(t.amount for t in transactions if t.amount)
            
            # Store in summary
            if market in summary:
                summary[market]["cost"] = abs(total_amount)
                summary[market]["charges"] = total_amount / len(transactions) if transactions else 0
                # Note: Losses % would come from parsed metadata if available
                summary[market]["losses_pct"] = 3.43  # Placeholder, should be from file
        
        return summary
    
    def get_sch_timeslot_data(
        self,
        sch_files: List[DailyFile],
        db: Session
    ) -> List[Dict]:
        """
        Extract 96 time-slot data from SCH files for Excel rows 7-102
        
        Args:
            sch_files: List of SCH file records
            db: Database session
            
        Returns:
            List of 96 time slots with consumption data
        """
        # Get all transactions ordered by time slot
        timeslot_data = []
        
        for sch_file in sch_files:
            transactions = db.query(Transaction).filter(
                Transaction.daily_file_id == sch_file.id
            ).order_by(Transaction.time_slot).all()
            
            for txn in transactions:
                timeslot_data.append({
                    "time_slot": txn.time_slot,
                    "consumption": txn.quantity_mw or 0,
                    "market": sch_file.market_type
                })
        
        return timeslot_data
    
    def populate_excel_with_data(
        self,
        excel_path: Path,
        dor_summary: Dict,
        sch_timeslots: List[Dict],
        calculation_date: date
    ) -> bool:
        """
        Populate Excel template with DOR and SCH data
        
        Args:
            excel_path: Path to monthly Excel file
            dor_summary: DOR summary data (GDAM, DAM, RTM)
            sch_timeslots: List of 96 time slots with consumption
            calculation_date: Date for this calculation
            
        Returns:
            True if successful
        """
        try:
            # Load workbook
            wb = load_workbook(excel_path)
            ws = wb.active
            
            print(f"📝 Populating Excel with data for {calculation_date}...")
            
            # Find the column for this date (assuming date headers in row 2)
            # For now, we'll use a simple column mapping
            # TODO: Make this dynamic based on actual template structure
            date_col = 3  # Column C (adjust based on template)
            
            # Row 4: GDAM data
            ws.cell(row=4, column=date_col, value=dor_summary["GDAM"]["losses_pct"])
            ws.cell(row=4, column=date_col+1, value=dor_summary["GDAM"]["charges"])
            ws.cell(row=4, column=date_col+2, value=dor_summary["GDAM"]["cost"])
            
            # Row 5: DAM data
            ws.cell(row=5, column=date_col, value=dor_summary["DAM"]["losses_pct"])
            ws.cell(row=5, column=date_col+1, value=dor_summary["DAM"]["charges"])
            ws.cell(row=5, column=date_col+2, value=dor_summary["DAM"]["cost"])
            
            # Row 6: RTM data
            ws.cell(row=6, column=date_col, value=dor_summary["RTM"]["losses_pct"])
            ws.cell(row=6, column=date_col+1, value=dor_summary["RTM"]["charges"])
            ws.cell(row=6, column=date_col+2, value=dor_summary["RTM"]["cost"])
            
            # Rows 7-102: SCH timeslot data (96 slots)
            for idx, slot_data in enumerate(sch_timeslots[:96], start=7):
                ws.cell(row=idx, column=date_col, value=slot_data.get("consumption", 0))
            
            # Save workbook
            wb.save(excel_path)
            print(f"✅ Excel populated successfully")
            
            return True
            
        except Exception as e:
            print(f"❌ Error populating Excel: {str(e)}")
            return False
    
    def read_calculated_results(
        self,
        excel_path: Path,
        calculation_date: date
    ) -> Dict:
        """
        Read calculated results from Excel (after formulas have executed)
        
        Args:
            excel_path: Path to monthly Excel file
            calculation_date: Date for this calculation
            
        Returns:
            Dictionary with calculated results
        """
        try:
            wb = load_workbook(excel_path, data_only=True)  # data_only=True to get calculated values
            ws = wb.active
            
            date_col = 3  # Column C (adjust based on template)
            
            # Read summary results (adjust rows based on actual template)
            results = {
                "calculation_date": str(calculation_date),
                "total_scheduled_kwh": 0,
                "total_consumption_kwh": 0,
                "total_losses_kwh": 0,
                "total_cost_inr": 0,
                "energy_savings_kwh": 0,
                "cost_savings_inr": 0,
                "hourly_data": []
            }
            
            # Read summary cells (assuming specific locations in template)
            # Row 103: Total Scheduled
            total_scheduled = ws.cell(row=103, column=date_col).value or 0
            results["total_scheduled_kwh"] = float(total_scheduled)
            
            # Row 104: Total after losses
            total_after_losses = ws.cell(row=104, column=date_col).value or 0
            results["total_consumption_kwh"] = float(total_after_losses)
            
            # Row 105: Total losses
            total_losses = ws.cell(row=105, column=date_col).value or 0
            results["total_losses_kwh"] = float(total_losses)
            
            # Row 106: Total cost
            total_cost = ws.cell(row=106, column=date_col).value or 0
            results["total_cost_inr"] = float(total_cost)
            
            # Row 107: Energy savings (if applicable)
            energy_savings = ws.cell(row=107, column=date_col).value or 0
            results["energy_savings_kwh"] = float(energy_savings)
            
            # Row 108: Cost savings
            cost_savings = ws.cell(row=108, column=date_col).value or 0
            results["cost_savings_inr"] = float(cost_savings)
            
            # Read hourly aggregated data (rows 109-132 for 24 hours)
            for hour in range(24):
                hour_row = 109 + hour
                hourly_consumption = ws.cell(row=hour_row, column=date_col).value or 0
                
                results["hourly_data"].append({
                    "hour": hour,
                    "consumption_kwh": float(hourly_consumption)
                })
            
            print(f"✅ Read calculated results: Savings = {results['energy_savings_kwh']:.2f} kWh")
            
            return results
            
        except Exception as e:
            print(f"❌ Error reading results: {str(e)}")
            return None
    
    def store_calculation_results(
        self,
        results: Dict,
        calculation_date: date,
        excel_path: Path,
        db: Session
    ) -> int:
        """
        Store calculated results in database
        
        Args:
            results: Calculated results from Excel
            calculation_date: Date for this calculation
            excel_path: Path to Excel file
            db: Database session
            
        Returns:
            ID of created daily record
        """
        try:
            # Create daily summary record
            daily_record = EnergyScheduleDaily(
                calculation_date=calculation_date,
                year=calculation_date.year,
                month=calculation_date.month,
                total_scheduled_kwh=results["total_scheduled_kwh"],
                total_consumption_kwh=results["total_consumption_kwh"],
                total_losses_kwh=results["total_losses_kwh"],
                total_cost_inr=results["total_cost_inr"],
                energy_savings_kwh=results["energy_savings_kwh"],
                cost_savings_inr=results["cost_savings_inr"],
                excel_file_path=str(excel_path),
                status="completed"
            )
            
            db.add(daily_record)
            db.flush()  # Get the ID
            
            # Create hourly records
            for hourly in results["hourly_data"]:
                hourly_record = EnergyScheduleHourly(
                    daily_id=daily_record.id,
                    hour=hourly["hour"],
                    consumption_kwh=hourly["consumption_kwh"]
                )
                db.add(hourly_record)
            
            db.commit()
            
            print(f"✅ Stored results in database (daily_id: {daily_record.id})")
            
            return daily_record.id
            
        except Exception as e:
            db.rollback()
            print(f"❌ Error storing results: {str(e)}")
            raise
    
    def calculate_energy_schedule(
        self,
        calculation_date: date,
        db: Session
    ) -> Dict:
        """
        Complete calculation workflow: validate → populate Excel → read results → store
        
        Args:
            calculation_date: Date for calculation
            db: Database session
            
        Returns:
            Dictionary with calculation results and status
        """
        print(f"\n{'='*80}")
        print(f"ENERGY SCHEDULE CALCULATION - {calculation_date}")
        print(f"{'='*80}\n")
        
        # Step 1: Validate files
        print("Step 1: Validating required files...")
        validation = self.validate_files_for_calculation(calculation_date, db)
        
        if validation["status"] != "ready":
            print(f"❌ Validation failed: {validation['message']}")
            return {
                "success": False,
                "message": validation["message"],
                "validation": validation
            }
        
        print(f"✅ Validation passed: {validation['message']}\n")
        
        # Step 2: Get or create monthly Excel
        print("Step 2: Preparing Excel file...")
        excel_path = self.get_or_create_monthly_excel(calculation_date)
        print()
        
        # Step 3: Extract data from database
        print("Step 3: Extracting data from database...")
        dor_date = calculation_date - timedelta(days=1)
        
        dor_files = db.query(DailyFile).filter(
            and_(
                DailyFile.trading_date == dor_date,
                DailyFile.report_type.like('DOR%')
            )
        ).all()
        
        sch_files = db.query(DailyFile).filter(
            and_(
                DailyFile.trading_date == calculation_date,
                DailyFile.report_type.like('SCH%')
            )
        ).all()
        
        dor_summary = self.get_dor_summary_data(dor_files, db)
        sch_timeslots = self.get_sch_timeslot_data(sch_files, db)
        
        print(f"  ✓ Extracted DOR summary for {len(dor_files)} files")
        print(f"  ✓ Extracted {len(sch_timeslots)} SCH timeslots\n")
        
        # Step 4: Populate Excel with data
        print("Step 4: Populating Excel with data...")
        success = self.populate_excel_with_data(
            excel_path, dor_summary, sch_timeslots, calculation_date
        )
        
        if not success:
            return {
                "success": False,
                "message": "Failed to populate Excel file"
            }
        
        print()
        
        # Step 5: Read calculated results
        print("Step 5: Reading calculated results...")
        results = self.read_calculated_results(excel_path, calculation_date)
        
        if not results:
            return {
                "success": False,
                "message": "Failed to read calculated results"
            }
        
        print()
        
        # Step 6: Store results in database
        print("Step 6: Storing results in database...")
        daily_id = self.store_calculation_results(
            results, calculation_date, excel_path, db
        )
        
        print(f"\n{'='*80}")
        print(f"✅ CALCULATION COMPLETE")
        print(f"{'='*80}")
        print(f"Date: {calculation_date}")
        print(f"Energy Savings: {results['energy_savings_kwh']:.2f} kWh")
        print(f"Cost Savings: ₹{results['cost_savings_inr']:.2f}")
        print(f"Excel File: {excel_path}")
        print(f"Database ID: {daily_id}")
        print(f"{'='*80}\n")
        
        return {
            "success": True,
            "message": "Calculation completed successfully",
            "daily_id": daily_id,
            "results": results,
            "excel_path": str(excel_path)
        }


# Singleton instance
calculator = EnergyScheduleCalculator()
