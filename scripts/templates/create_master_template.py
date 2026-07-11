"""Create master energy schedule Excel template"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

def create_master_template():
    """Create the master energy schedule template Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Energy Schedule"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header section
    ws['A1'] = "ENERGY SCHEDULE - MONTHLY REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = "Portfolio:"
    ws['B2'] = "[PORTFOLIO_CODE]"  # Placeholder
    ws['A3'] = "Month:"
    ws['B3'] = "[MONTH_YEAR]"  # Placeholder
    ws['A4'] = "Generated:"
    ws['B4'] = "[GENERATION_DATE]"  # Placeholder
    
    # Column headers (row 6)
    headers = ['Date', 'Time Slot', 'SCH After Loss (MW)', 'GDAM Charges (₹)', 
               'DAM Charges (₹)', 'RTM Charges (₹)', 'Total Charges (₹)', 'Net Position']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Set column widths
    column_widths = [12, 15, 18, 18, 18, 18, 18, 15]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Data section starts at row 7
    # For a full month (31 days × 96 timeslots = 2976 rows max)
    # We'll create placeholders for first day as example
    
    start_row = 7
    example_date = "2026-01-01"
    
    # Add example data structure (first 5 timeslots)
    time_slots = [
        "00:00-00:15", "00:15-00:30", "00:30-00:45", "00:45-01:00", "01:00-01:15"
    ]
    
    for idx, time_slot in enumerate(time_slots):
        row = start_row + idx
        ws.cell(row=row, column=1).value = example_date  # Date
        ws.cell(row=row, column=2).value = time_slot  # Time Slot
        ws.cell(row=row, column=3).value = "[SCH_AFTER_LOSS]"  # SCH After Loss
        ws.cell(row=row, column=4).value = "[GDAM_CHARGE]"  # GDAM Charges
        ws.cell(row=row, column=5).value = "[DAM_CHARGE]"  # DAM Charges
        ws.cell(row=row, column=6).value = "[RTM_CHARGE]"  # RTM Charges
        ws.cell(row=row, column=7).value = f"=D{row}+E{row}+F{row}"  # Total formula
        ws.cell(row=row, column=8).value = f"=C{row}-G{row}"  # Net Position formula
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
            if col >= 3:  # Numeric columns
                ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Summary section (example at row 13)
    summary_row = start_row + len(time_slots) + 2
    ws.cell(row=summary_row, column=1).value = "MONTHLY SUMMARY"
    ws.cell(row=summary_row, column=1).font = subheader_font
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    
    summary_labels = [
        ("Total SCH After Loss:", "=SUM(C7:C2976)"),
        ("Total GDAM Charges:", "=SUM(D7:D2976)"),
        ("Total DAM Charges:", "=SUM(E7:E2976)"),
        ("Total RTM Charges:", "=SUM(F7:F2976)"),
        ("Total All Charges:", "=SUM(G7:G2976)"),
        ("Net Monthly Position:", "=SUM(H7:H2976)")
    ]
    
    for idx, (label, formula) in enumerate(summary_labels):
        row = summary_row + idx + 1
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=1).font = subheader_font
        ws.cell(row=row, column=2).value = formula
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=2).border = border
    
    # Instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        "ENERGY SCHEDULE TEMPLATE - USAGE INSTRUCTIONS",
        "",
        "This template is used for monthly energy schedule calculations.",
        "",
        "DATA MAPPING:",
        "1. SCH After Loss: Populated from SCH files (scheduled energy after transmission loss)",
        "2. GDAM/DAM/RTM Charges: Populated from DOR files (charges for each market type)",
        "",
        "WORKFLOW:",
        "1. System copies this template to calculations/YYYY-MM/ folder",
        "2. Populates data for all days in the month (up to 31 days × 96 timeslots)",
        "3. Formulas auto-calculate totals and net positions",
        "4. Results are read back and stored in database",
        "",
        "PLACEHOLDERS:",
        "[PORTFOLIO_CODE] - Replaced with actual portfolio code",
        "[MONTH_YEAR] - Replaced with month and year (e.g., January 2026)",
        "[GENERATION_DATE] - Replaced with calculation timestamp",
        "[SCH_AFTER_LOSS] - Replaced with actual SCH quantity",
        "[GDAM_CHARGE/DAM_CHARGE/RTM_CHARGE] - Replaced with actual charges",
        "",
        "DO NOT MODIFY:",
        "- Column structure (A-H)",
        "- Formula cells in columns G and H",
        "- Summary formulas in summary section"
    ]
    
    for idx, instruction in enumerate(instructions, start=1):
        ws_instructions.cell(row=idx, column=1).value = instruction
        if idx == 1:
            ws_instructions.cell(row=idx, column=1).font = Font(bold=True, size=12)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    # Save template
    templates_dir = Path("templates")
    templates_dir.mkdir(exist_ok=True)
    
    template_path = templates_dir / "master_energy_schedule.xlsx"
    wb.save(template_path)
    
    print(f"✅ Master template created: {template_path}")
    print(f"   - Main sheet: 'Energy Schedule' with data structure")
    print(f"   - Instructions sheet: Usage guidelines")
    print(f"   - Ready for monthly calculations")

if __name__ == "__main__":
    create_master_template()
