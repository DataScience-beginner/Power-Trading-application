#!/usr/bin/env python3
"""
Create proper Excel template matching the screenshot
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = 'Energy Schedule'

# Row 1: Dates (will be filled dynamically)
ws['D1'] = '01-12-2022'
ws['G1'] = '02-12-2022'
ws['J1'] = '03-12-2022'
ws['M1'] = '04-12-2022'

# Row 2: Market types
ws['C2'] = 'GDAM'
ws['D2'] = 'DAM'
ws['E2'] = 'RTM'
ws['F2'] = 'GDAM'
ws['G2'] = 'DAM'
ws['H2'] = 'RTM'
ws['I2'] = 'GDAM'
ws['J2'] = 'DAM'
ws['K2'] = 'RTM'
ws['L2'] = 'GDAM'
ws['M2'] = 'DAM'
ws['N2'] = 'RTM'

# Row 3: CTU Losses (will be calculated)
ws['A3'] = 'CTU Losses'
ws['A3'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Row 4: CTU Charges
ws['A4'] = 'CTU Charges'
ws['A4'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Row 5: Cost
ws['A5'] = 'Cost'
ws['A5'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Row 6: NLDC App Fee
ws['A6'] = 'NLDC App Fee'
ws['A6'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Timeslots (96 slots = 15-minute intervals)
timeslots = []
for hour in range(24):
    for minute in [0, 15, 30, 45]:
        start_time = f"{hour:02d}:{minute:02d}"
        end_minute = minute + 15
        end_hour = hour
        if end_minute >= 60:
            end_minute = 0
            end_hour += 1
        end_time = f"{end_hour:02d}:{end_minute:02d}"
        timeslots.append(f"{start_time} - {end_time}")

# Add timeslots starting from row 7
for idx, slot in enumerate(timeslots, start=7):
    ws[f'A{idx}'] = slot

wb.save('templates/master_energy_schedule.xlsx')
print('✅ Created proper Excel template with timeslots')
