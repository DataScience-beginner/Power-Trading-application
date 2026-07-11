#!/usr/bin/env python3
"""
Generate 15 Days of Realistic Mock Trading Data
===============================================

Creates mock DOR and SCH files for Jan 2-16, 2026 with:
- 5 different clients
- Realistic trading patterns (not all markets every day)
- Proper paired DOR+SCH files for energy schedule calculation
- Some days with only 2-3 files (more realistic)
"""

import os
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import random
import openpyxl

# Configuration
START_DATE = datetime(2026, 1, 2)  # Jan 2, 2026
NUM_DAYS = 15
OUTPUT_DIR = Path("Data/mock_reports_15days")

# 5 Clients with their details
CLIENTS = [
    {
        "entity_id": "A2AR0NPT0001",
        "entity_name": "Tata Power Company Limited",
        "portfolio_code": "NPT0001_TN0",
        "state": "Tamil Nadu"
    },
    {
        "entity_id": "A2AR0NPT0002",
        "entity_name": "Adani Power Limited",
        "portfolio_code": "NPT0002_GJ0",
        "state": "Gujarat"
    },
    {
        "entity_id": "A2AR0NPT0003",
        "entity_name": "Tamil Nadu Generation and Distribution Corporation",
        "portfolio_code": "NPT0003_TN0",
        "state": "Tamil Nadu"
    },
    {
        "entity_id": "A2AR0NPT0004",
        "entity_name": "BSES Rajdhani Power Limited",
        "portfolio_code": "NPT0004_DL0",
        "state": "Delhi"
    },
    {
        "entity_id": "A2AR0NPT0005",
        "entity_name": "Bangalore Electricity Supply Company",
        "portfolio_code": "NPT0005_KA0",
        "state": "Karnataka"
    }
]

# Market types - not all clients trade in all markets every day
MARKET_TYPES = ["GDAM", "DAM", "RTM"]

# Template files
TEMPLATE_DIR = Path("Data/templates")
DOR_TEMPLATE = TEMPLATE_DIR / "GDAM_template_DOR.xls"
SCH_TEMPLATE = TEMPLATE_DIR / "GDAM_template_SCH.xls"


def create_template_directory():
    """Create template files if they don't exist"""
    TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Check if we have existing sample files to use as templates
    mock_reports = Path("Data/mock_reports")
    if mock_reports.exists():
        # Find first DOR and SCH file
        dor_files = list(mock_reports.glob("*DOR*.xls*"))
        sch_files = list(mock_reports.glob("*SCH*.xls*"))
        
        if dor_files and not DOR_TEMPLATE.exists():
            shutil.copy(dor_files[0], DOR_TEMPLATE)
            print(f"✅ Created DOR template from {dor_files[0].name}")
        
        if sch_files and not SCH_TEMPLATE.exists():
            shutil.copy(sch_files[0], SCH_TEMPLATE)
            print(f"✅ Created SCH template from {sch_files[0].name}")
    
    return DOR_TEMPLATE.exists() and SCH_TEMPLATE.exists()


def should_trade_in_market(client_idx, market_type, day_of_week):
    """Determine if a client trades in a specific market on a given day
    
    Makes it more realistic - not all clients trade in all markets every day
    """
    
    # GDAM is most common - 90% participation
    if market_type == "GDAM":
        return random.random() < 0.95
    
    # DAM is common - 70% participation
    elif market_type == "DAM":
        return random.random() < 0.80
    
    # RTM is less common - 40% participation (intra-day adjustments)
    elif market_type == "RTM":
        # Less RTM trading on weekends
        if day_of_week >= 5:  # Saturday/Sunday
            return random.random() < 0.20
        return random.random() < 0.50
    
    return True


def modify_excel_cells(file_path, client, market_type, trading_date):
    """Modify Excel file cells with client-specific data"""
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Update client information in known cell locations
        # These are typical locations for DOR/SCH reports
        
        # Entity ID (usually in top rows)
        for row in range(1, 15):
            for col in range(1, 8):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    # Replace entity ID patterns
                    if "NPT0" in cell_value or "A2AR" in cell_value:
                        ws.cell(row, col).value = cell_value.replace(
                            ws.cell(row, col).value.split('_')[0], 
                            client['entity_id']
                        )
                    
                    # Replace company name
                    if "Company Limited" in cell_value or "Corporation" in cell_value:
                        ws.cell(row, col).value = client['entity_name']
                    
                    # Replace portfolio code
                    if "_TN0" in cell_value or "_KA0" in cell_value or "_GJ0" in cell_value or "_DL0" in cell_value:
                        old_portfolio = cell_value
                        new_portfolio = old_portfolio.split('_')[0] + '_' + client['portfolio_code'].split('_')[1]
                        ws.cell(row, col).value = cell_value.replace(old_portfolio, client['portfolio_code'])
        
        # Update date if found
        for row in range(1, 10):
            for col in range(1, 8):
                cell_value = ws.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if "Date" in cell_value or "Trading Day" in cell_value:
                        # Update date in next cell
                        ws.cell(row, col + 1).value = trading_date.strftime("%d-%b-%Y")
        
        # Save modified workbook
        wb.save(file_path)
        return True
        
    except Exception as e:
        print(f"⚠️  Error modifying Excel cells: {e}")
        return False


def generate_realistic_15days():
    """Generate 15 days of realistic mock data"""
    
    print("\n" + "="*70)
    print("  Generate 15 Days of Realistic Trading Data")
    print("="*70 + "\n")
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create templates
    if not create_template_directory():
        print("❌ Template files not found!")
        print("   Please ensure Data/mock_reports/ contains sample DOR and SCH files")
        return
    
    total_files = 0
    file_breakdown = {
        "GDAM_DOR": 0, "DAM_DOR": 0, "RTM_DOR": 0,
        "GDAM_SCH": 0, "DAM_SCH": 0, "RTM_SCH": 0
    }
    
    # Generate data for each day
    for day_offset in range(NUM_DAYS):
        trading_date = START_DATE + timedelta(days=day_offset)
        day_of_week = trading_date.weekday()
        
        print(f"\n📅 {trading_date.strftime('%Y-%m-%d (%A)')}")
        print("-" * 70)
        
        day_files = 0
        
        # Generate for each client
        for client_idx, client in enumerate(CLIENTS):
            client_files = []
            
            # Determine which markets this client trades in today
            for market_type in MARKET_TYPES:
                if should_trade_in_market(client_idx, market_type, day_of_week):
                    # Create DOR file
                    dor_filename = f"{market_type}_IEX{trading_date.strftime('%d%m%y')}DOR_{client['entity_id']}_{client['portfolio_code']}_{client['entity_name'].replace(' ', '_')}.xlsx"
                    dor_path = OUTPUT_DIR / dor_filename
                    
                    # Copy from template
                    shutil.copy(DOR_TEMPLATE, dor_path)
                    
                    # Modify cells
                    if modify_excel_cells(dor_path, client, market_type, trading_date):
                        client_files.append(f"DOR-{market_type}")
                        total_files += 1
                        day_files += 1
                        file_breakdown[f"{market_type}_DOR"] += 1
                    
                    # Create corresponding SCH file (required for energy schedule)
                    # SCH is usually created for the NEXT day's schedule
                    sch_date = trading_date
                    sch_filename = f"{market_type}_IEX{sch_date.strftime('%d%m%y')}SCH_{client['entity_id']}_{client['portfolio_code']}_{client['entity_name'].replace(' ', '_')}.xlsx"
                    sch_path = OUTPUT_DIR / sch_filename
                    
                    # Copy from template
                    shutil.copy(SCH_TEMPLATE, sch_path)
                    
                    # Modify cells
                    if modify_excel_cells(sch_path, client, market_type, sch_date):
                        client_files.append(f"SCH-{market_type}")
                        total_files += 1
                        day_files += 1
                        file_breakdown[f"{market_type}_SCH"] += 1
            
            if client_files:
                print(f"   {client['entity_name'][:30]:30} → {', '.join(client_files)}")
        
        print(f"   Total files for this day: {day_files}")
    
    # Summary
    print("\n" + "="*70)
    print("  GENERATION COMPLETE")
    print("="*70)
    print(f"\n📊 Total Files Generated: {total_files}")
    print(f"   Output Directory: {OUTPUT_DIR}")
    print()
    print("File Breakdown:")
    for file_type, count in file_breakdown.items():
        print(f"   {file_type:12} → {count:3} files")
    print()
    print("Next Steps:")
    print("   1. Run: python upload_15day_data.py")
    print("   2. Check Railway dashboard for energy schedule calculations")
    print()


if __name__ == "__main__":
    random.seed(42)  # For reproducible results
    generate_realistic_15days()
