"""
Generate realistic mock GDAM/DAM/RTM DOR reports by copying real file structure
Creates proper IEX format files with different client names
"""

import openpyxl
from openpyxl import load_workbook
import shutil
from datetime import datetime, timedelta
from pathlib import Path
import random

# Configuration
NUM_DAYS = 30
START_DATE = datetime(2026, 1, 1)
OUTPUT_DIR = Path("Data/mock_reports_real_format")
TEMPLATE_FILE = "Data/GDAM_IEX120126DOR_NPT0027_KA0_Mellbro_Sugars_Pvt.xls"

# Different clients for variety
CLIENTS = [
    {"code": "NPT0001_MH0", "name": "Tata_Power_Limited", "entity": "Tata Power Company Limited"},
    {"code": "NPT0002_GJ0", "name": "Adani_Power_Limited", "entity": "Adani Power Limited"},
    {"code": "NPT0003_TN0", "name": "TANGEDCO_Chennai", "entity": "Tamil Nadu Generation and Distribution Corporation"},
    {"code": "NPT0004_DL0", "name": "BSES_Rajdhani", "entity": "BSES Rajdhani Power Limited"},
    {"code": "NPT0005_KA0", "name": "BESCOM_Bangalore", "entity": "Bangalore Electricity Supply Company"}
]

def convert_xls_to_xlsx(xls_path):
    """Convert .xls to .xlsx using LibreOffice"""
    import subprocess
    import tempfile
    import os
    
    temp_dir = tempfile.gettempdir()
    result = subprocess.run([
        'soffice', '--headless', '--convert-to', 'xlsx',
        '--outdir', temp_dir, xls_path
    ], capture_output=True, timeout=30)
    
    if result.returncode == 0:
        base_name = os.path.splitext(os.path.basename(xls_path))[0]
        xlsx_path = os.path.join(temp_dir, f"{base_name}.xlsx")
        if os.path.exists(xlsx_path):
            return xlsx_path
    raise Exception("LibreOffice conversion failed")

def modify_dates_in_workbook(wb, trading_date, delivery_date):
    """Modify dates in the workbook"""
    ws = wb.active
    
    # Find and update dates (they're typically in the first 20 rows)
    trading_str = trading_date.strftime("%d.%m.%Y")
    delivery_str = delivery_date.strftime("%d.%m.%Y")
    
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if cell.value:
                val = str(cell.value)
                # Look for date patterns
                if '12.01.2026' in val:
                    cell.value = val.replace('12.01.2026', trading_str)
                elif '13.01.2026' in val:
                    cell.value = val.replace('13.01.2026', delivery_str)

def randomize_quantities(wb):
    """Randomize transaction quantities while keeping structure"""
    ws = wb.active
    
    # Find the data table (usually starts after row 10)
    for row in ws.iter_rows(min_row=10):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                # Add some random variation (+/- 20%)
                variation = random.uniform(0.8, 1.2)
                cell.value = round(cell.value * variation, 2)

def create_mock_file(template_xlsx, output_path, trading_date, delivery_date, client):
    """Create a mock file by modifying template"""
    
    # Load template
    wb = load_workbook(template_xlsx)
    
    # Modify dates
    modify_dates_in_workbook(wb, trading_date, delivery_date)
    
    # Randomize quantities
    randomize_quantities(wb)
    
    # TODO: Update client name in workbook (requires finding exact cell location)
    
    # Save
    wb.save(output_path)
    wb.close()

def main():
    print("="*70)
    print("MOCK DATA GENERATOR - Real IEX Format")
    print("="*70)
    
    # Create output directory
    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"✓ Output directory: {OUTPUT_DIR}")
    
    # Convert template to xlsx first
    print(f"\n📄 Converting template file...")
    try:
        template_xlsx = convert_xls_to_xlsx(TEMPLATE_FILE)
        print(f"✓ Template converted: {template_xlsx}")
    except Exception as e:
        print(f"❌ Error: {e}")
        return
    
    print(f"\n🔧 Generating {NUM_DAYS} days × 3 markets = {NUM_DAYS * 3} files...")
    print(f"📅 Date range: {START_DATE.strftime('%Y-%m-%d')} to {(START_DATE + timedelta(days=NUM_DAYS-1)).strftime('%Y-%m-%d')}\n")
    
    success_count = 0
    
    for day_num in range(NUM_DAYS):
        trading_date = START_DATE + timedelta(days=day_num)
        delivery_date = trading_date + timedelta(days=1)
        
        # Rotate through clients
        client = CLIENTS[day_num % len(CLIENTS)]
        
        date_str = trading_date.strftime("%d%m%y")
        
        print(f"Day {day_num+1}/{NUM_DAYS}: {trading_date.strftime('%Y-%m-%d')} - {client['name']}")
        
        # Generate GDAM, DAM, RTM files
        for market in ['GDAM', 'DAM', 'RTM']:
            filename = f"{market}_IEX{date_str}DOR_{client['code']}_{client['name']}.xlsx"
            output_path = OUTPUT_DIR / filename
            
            try:
                create_mock_file(template_xlsx, output_path, trading_date, delivery_date, client)
                print(f"  ✓ {market}")
                success_count += 1
            except Exception as e:
                print(f"  ❌ {market}: {e}")
    
    print(f"\n{'='*70}")
    print(f"✅ Generated {success_count} mock files")
    print(f"📁 Location: {OUTPUT_DIR}")
    print(f"{'='*70}")
    print("\nNEXT STEPS:")
    print("1. Test locally: python run_parser.py 'Data/mock_reports_real_format/GDAM_*.xlsx'")
    print("2. Upload to Railway: python upload_to_railway_fast.py")
    print(f"{'='*70}")

if __name__ == "__main__":
    main()
