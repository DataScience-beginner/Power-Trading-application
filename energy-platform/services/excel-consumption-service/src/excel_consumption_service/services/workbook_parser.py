from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


MONTH_LOOKUP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}

DAM_RTM_CATEGORY_ROWS = {
    "C1": 100,
    "C2": 101,
    "C4": 103,
    "C5": 104,
}

TNEB_CATEGORY_COLUMNS = {
    "C1": 9,
    "C2": 12,
    "C4": 18,
    "C5": 21,
}


@dataclass(slots=True)
class ParsedSheetSummaryData:
    sheet_name: str
    sheet_type: str
    status: str
    row_count: int | None = None
    validation_summary: str | None = None


@dataclass(slots=True)
class ParsedIntervalRecordData:
    source_type: str
    reading_date: date
    time_block_label: str
    category_code: str
    quantity: float


@dataclass(slots=True)
class ParsedDailyValueData:
    source_type: str
    reading_date: date
    metric_name: str
    category_code: str | None
    value: float


@dataclass(slots=True)
class ParsedWorkbookData:
    workbook_month: str | None
    sheet_summaries: list[ParsedSheetSummaryData] = field(default_factory=list)
    interval_records: list[ParsedIntervalRecordData] = field(default_factory=list)
    daily_values: list[ParsedDailyValueData] = field(default_factory=list)
    solar_values: list[ParsedDailyValueData] = field(default_factory=list)


def parse_workbook(file_name: str, content: bytes) -> ParsedWorkbookData:
    """Parse the known workbook format into normalized in-memory records."""

    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    required_sheets = {"DAM", "RTM", "TNEB"}
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise ValueError(
            f"Workbook is missing required sheets: {', '.join(sorted(missing))}."
        )

    tneb_values = _parse_tneb_sheet(workbook["TNEB"])
    legacy_solar_values = (
        _parse_legacy_solar_working(workbook["Solar Working"])
        if "Solar Working" in workbook.sheetnames
        else []
    )
    workbook_period = _infer_workbook_period(
        file_name=file_name,
        tneb_dates=[record.reading_date for record in tneb_values],
        solar_dates=[record.reading_date for record in legacy_solar_values],
    )

    parsed = ParsedWorkbookData(
        workbook_month=workbook_period.strftime("%Y-%m") if workbook_period else None,
    )
    parsed.daily_values.extend(_filter_records_to_workbook_month(tneb_values, workbook_period))
    parsed.solar_values.extend(
        _filter_records_to_workbook_month(legacy_solar_values, workbook_period)
    )
    parsed.sheet_summaries.append(
        ParsedSheetSummaryData(
            sheet_name="TNEB",
            sheet_type="daily-consumption",
            status="parsed",
            row_count=workbook["TNEB"].max_row,
            validation_summary=f"Parsed {len(tneb_values)} normalized daily values.",
        )
    )

    if legacy_solar_values:
        parsed.sheet_summaries.append(
            ParsedSheetSummaryData(
                sheet_name="Solar Working",
                sheet_type="legacy-solar-source",
                status="parsed",
                row_count=workbook["Solar Working"].max_row,
                validation_summary="Used legacy Solar Working totals as the temporary solar source.",
            )
        )

    for sheet_name in ("DAM", "RTM"):
        interval_values, daily_values = _parse_market_sheet(
            worksheet=workbook[sheet_name],
            source_type=sheet_name,
            workbook_period=workbook_period,
        )
        parsed.interval_records.extend(interval_values)
        parsed.daily_values.extend(daily_values)
        parsed.sheet_summaries.append(
            ParsedSheetSummaryData(
                sheet_name=sheet_name,
                sheet_type="market-source",
                status="parsed",
                row_count=workbook[sheet_name].max_row,
                validation_summary=(
                    f"Parsed {len(interval_values)} interval records and "
                    f"{len(daily_values)} category totals."
                ),
            )
        )

    if "Solar" in workbook.sheetnames:
        parsed.sheet_summaries.append(
            ParsedSheetSummaryData(
                sheet_name="Solar",
                sheet_type="solar-source",
                status="recognized",
                row_count=workbook["Solar"].max_row,
                validation_summary="Dedicated Solar sheet detected. Parsing rules are reserved for the next iteration.",
            )
        )

    return parsed


def _parse_tneb_sheet(worksheet) -> list[ParsedDailyValueData]:
    records: list[ParsedDailyValueData] = []
    for row_index in range(12, worksheet.max_row + 1):
        reading_date = _coerce_excel_date(worksheet.cell(row_index, 1).value)
        if reading_date is None:
            continue

        total_value = _coerce_number(worksheet.cell(row_index, 3).value)
        if total_value is not None:
            records.append(
                ParsedDailyValueData(
                    source_type="TNEB",
                    reading_date=reading_date,
                    metric_name="tneb_total_consumption",
                    category_code="TOTAL",
                    value=total_value,
                )
            )

        for category_code, column_index in TNEB_CATEGORY_COLUMNS.items():
            category_value = _coerce_number(worksheet.cell(row_index, column_index).value)
            if category_value is None:
                continue
            records.append(
                ParsedDailyValueData(
                    source_type="TNEB",
                    reading_date=reading_date,
                    metric_name="tneb_category_total",
                    category_code=category_code,
                    value=category_value,
                )
            )
    return records


def _parse_legacy_solar_working(worksheet) -> list[ParsedDailyValueData]:
    records: list[ParsedDailyValueData] = []
    for row_index in range(4, worksheet.max_row + 1):
        reading_date = _coerce_excel_date(worksheet.cell(row_index, 2).value)
        if reading_date is None:
            continue

        solar_total = _coerce_number(worksheet.cell(row_index, 18).value)
        if solar_total is None:
            continue
        records.append(
            ParsedDailyValueData(
                source_type="SOLAR_WORKING_LEGACY",
                reading_date=reading_date,
                metric_name="legacy_solar_total",
                category_code="TOTAL",
                value=solar_total,
            )
        )
    return records


def _parse_market_sheet(
    worksheet,
    source_type: str,
    workbook_period: date | None,
) -> tuple[list[ParsedIntervalRecordData], list[ParsedDailyValueData]]:
    interval_records: list[ParsedIntervalRecordData] = []
    daily_values: list[ParsedDailyValueData] = []

    if workbook_period is not None:
        for row_index in range(3, 99):
            time_block_label = worksheet.cell(row_index, 1).value
            if not time_block_label:
                continue
            for column_index in range(2, 33):
                day_number = _coerce_day_number(worksheet.cell(2, column_index).value)
                if day_number is None:
                    continue
                reading_date = _build_date(workbook_period, day_number)
                if reading_date is None:
                    continue
                quantity = _coerce_number(worksheet.cell(row_index, column_index).value)
                if quantity in (None, 0):
                    continue
                interval_records.append(
                    ParsedIntervalRecordData(
                        source_type=source_type,
                        reading_date=reading_date,
                        time_block_label=str(time_block_label),
                        category_code="TOTAL",
                        quantity=quantity,
                    )
                )

        for category_code, row_index in DAM_RTM_CATEGORY_ROWS.items():
            for column_index in range(2, 33):
                day_number = _coerce_day_number(worksheet.cell(2, column_index).value)
                if day_number is None:
                    continue
                reading_date = _build_date(workbook_period, day_number)
                if reading_date is None:
                    continue
                value = _coerce_number(worksheet.cell(row_index, column_index).value)
                if value is None:
                    continue
                daily_values.append(
                    ParsedDailyValueData(
                        source_type=source_type,
                        reading_date=reading_date,
                        metric_name="iex_category_total",
                        category_code=category_code,
                        value=value,
                    )
                )

    return interval_records, daily_values


def _infer_workbook_period(
    file_name: str,
    tneb_dates: list[date],
    solar_dates: list[date],
) -> date | None:
    candidate_dates = solar_dates + tneb_dates
    if candidate_dates:
        year, month = Counter(
            (item.year, item.month) for item in candidate_dates
        ).most_common(1)[0][0]
        return date(year, month, 1)

    match = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)['\\s_-]*(\\d{2,4})",
        Path(file_name).stem,
        re.IGNORECASE,
    )
    if not match:
        return None

    month = MONTH_LOOKUP[match.group(1).lower()]
    year_token = match.group(2)
    year = int(year_token)
    if year < 100:
        year += 2000
    return date(year, month, 1)


def _build_date(period_start: date, day_number: int) -> date | None:
    try:
        return date(period_start.year, period_start.month, day_number)
    except ValueError:
        return None


def _coerce_excel_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 30000:
        return from_excel(value).date()
    return None


def _coerce_number(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not cleaned:
            return None
        return float(cleaned)
    return None


def _coerce_day_number(value) -> int | None:
    if isinstance(value, int):
        return value if 1 <= value <= 31 else None
    if isinstance(value, float) and value.is_integer():
        integer_value = int(value)
        return integer_value if 1 <= integer_value <= 31 else None
    return None


def _filter_records_to_workbook_month(records, workbook_period: date | None):
    if workbook_period is None:
        return records
    return [
        record
        for record in records
        if (
            record.reading_date.year == workbook_period.year
            and record.reading_date.month == workbook_period.month
        )
    ]
